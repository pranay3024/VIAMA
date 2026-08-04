"""
VIAMA - every JSON API route.

Mounted under the configured base path (default /api/v1) by
core.api.register_api().
"""

from flask import Blueprint as _Blueprint

#: One blueprint carries every route below.  Each section used to declare its
#: own; after consolidation they share this one.  The URLs are unchanged - only
#: how they get registered.
bp = _Blueprint("viama_api", __name__)

# ==========================================================================
# res_meta
# Auth and metadata endpoints.
#
# ``/meta/*`` exposes everything the portal hardcodes in Python or Jinja - week
# lists, the team->state map, status vocabularies, dropdown contents and both
# checklists - so the consuming site never has to duplicate a constant.
# ==========================================================================

import os

from flask import Blueprint, request

from core import config as checklist_data
from core.engine import (
    ALL_SCOPES,
    READ_ONLY_SCOPES,
    current_client,
    decode_token,
    mint_token,
    require_auth,
)
from core.config import (
    ALL_DAYS,
    ALL_SURVEY_TYPES,
    ASSIGNMENT_STATUSES,
    DEFECT_FIELD_META,
    EMAIL_TYPES,
    KNOWN_STATES,
    PROJECT_START,
    ROLES,
    SCHEDULE_DAYS,
    SURVEY_STATUSES,
    SURVEY_STATUS_LABELS,
    TASK_FIELDS,
    TEAMS,
    WEEKDAY_ORDER,
    team_display,
)
from core.engine import ok
from core.engine import BadRequest, Forbidden, NotFound, ValidationError
from core.config import ASSUMED_SEMANTICS, COLUMN_SEMANTICS, ist_now_aware, iso, utc_now
from core.config import current_week_number, week_detail, week_list



# ---------------------------------------------------------------------------
# Health / version  (no auth - used by uptime checks)
# ---------------------------------------------------------------------------


@bp.get("/meta/health")
def health():
    from extensions import db

    db_ok, db_error = True, None
    try:
        db.session.execute(db.text("SELECT 1"))
    except Exception as exc:  # pragma: no cover - depends on live DB
        db_ok = False
        db_error = type(exc).__name__

    payload = {
        "status": "ok" if db_ok else "degraded",
        "database": {"ok": db_ok, "error": db_error},
        "time_utc": iso(utc_now().replace(microsecond=0)) + "Z",
        "time_ist": iso(ist_now_aware().replace(microsecond=0)),
        "api_version": "v1",
    }
    return ok(payload, status=200 if db_ok else 503)


@bp.get("/openapi.json")
@require_auth()
def openapi_spec():
    """
    Machine-readable description of every route, generated from the live app.

    Not hand-maintained, so it cannot drift: it reflects exactly what is mounted
    right now. Feed it to a client generator on the other VM.
    """
    from flask import current_app

    from core.api import base_path

    mount = base_path()
    paths = {}

    for rule in current_app.url_map.iter_rules():
        if not str(rule).startswith(mount):
            continue

        path = str(rule)
        # Flask "<int:survey_id>" -> OpenAPI "{survey_id}"
        for converter in ("int:", "path:", "string:", "float:"):
            path = path.replace("<" + converter, "{")
        path = path.replace("<", "{").replace(">", "}")

        view = current_app.view_functions.get(rule.endpoint)
        doc = (view.__doc__ or "").strip() if view else ""
        summary = doc.splitlines()[0] if doc else rule.endpoint

        params = [
            {
                "name": arg,
                "in": "path",
                "required": True,
                "schema": {"type": "integer" if "int:" in str(rule) else "string"},
            }
            for arg in rule.arguments
        ]

        for method in sorted(rule.methods - {"HEAD", "OPTIONS"}):
            paths.setdefault(path, {})[method.lower()] = {
                "operationId": f"{rule.endpoint}_{method.lower()}",
                "summary": summary,
                "description": doc,
                "tags": [path[len(mount):].strip("/").split("/")[0] or "root"],
                "parameters": params,
                "security": [{"bearerAuth": []}],
                "responses": {
                    "200": {"description": "Success"},
                    "401": {"description": "Missing or invalid token"},
                    "403": {"description": "Token lacks the required scope"},
                    "404": {"description": "Not found"},
                    "409": {"description": "Conflicts with current state"},
                    "422": {"description": "Validation failed"},
                },
            }

    return ok(
        {
            "openapi": "3.0.3",
            "info": {
                "title": "VIAMA Surveillance Portal API",
                "version": "1.0.0",
                "description": (
                    "Read/write access to every record the portal holds, the "
                    "derived values its dashboards display, full-database dumps, "
                    "access logs, and a webhook + change-feed push channel."
                ),
            },
            "servers": [{"url": mount}],
            "components": {
                "securitySchemes": {
                    "bearerAuth": {
                        "type": "http",
                        "scheme": "bearer",
                        "bearerFormat": "JWT",
                    }
                }
            },
            "security": [{"bearerAuth": []}],
            "paths": paths,
            "x-route-count": len(paths),
        }
    )


@bp.get("/meta/endpoints")
@require_auth()
def list_endpoints_summary():
    """Compact route listing, grouped by area - easier to skim than OpenAPI."""
    from flask import current_app

    from core.api import base_path

    mount = base_path()
    groups = {}

    for rule in current_app.url_map.iter_rules():
        text = str(rule)
        if not text.startswith(mount):
            continue
        tail = text[len(mount):].strip("/")
        group = tail.split("/")[0] or "root"
        methods = sorted(rule.methods - {"HEAD", "OPTIONS"})
        groups.setdefault(group, []).append(
            {"path": text, "methods": methods}
        )

    for items in groups.values():
        items.sort(key=lambda item: item["path"])

    return ok(
        {
            "base_path": mount,
            "total": sum(len(v) for v in groups.values()),
            "groups": {k: groups[k] for k in sorted(groups)},
        }
    )


@bp.get("/meta/integrity")
def integrity():
    """
    Self-check: are the API's own files still present and importable?

    Unauthenticated on purpose so a monitor can poll it, and it deliberately
    reveals nothing beyond presence/absence.

    Poll this from the consuming VM. If a core file gets deleted, this goes
    ``ok: false`` immediately - rather than the integration failing silently and
    being noticed days later through missing data.
    """
    from core.api import check_integrity

    report = check_integrity()
    return ok(
        {
            "ok": report["ok"],
            "missing": report["missing"],
            "broken": report["broken"],
            "module_count": len(report["modules"]),
            "remedy": report.get("remedy"),
        },
        status=200 if report["ok"] else 503,
    )


@bp.get("/meta/version")
def version():
    """
    Version plus the timezone contract.

    ``datetime_semantics`` tells the consumer how to read every timestamp; see
    core/timeutils.py for why the columns disagree.
    """
    semantics = {}
    for (table, column), semantic in COLUMN_SEMANTICS.items():
        semantics.setdefault(table, {})[column] = {
            "semantic": semantic,
            "confidence": "assumed" if (table, column) in ASSUMED_SEMANTICS else "verified",
        }

    return ok(
        {
            "api_version": "v1",
            "project_start": PROJECT_START.isoformat(),
            "current_week": current_week_number(),
            "datetime_semantics": semantics,
            "notes": {
                "ist_wall": (
                    "Stored value is IST wall-clock; *_utc is raw minus 05:30."
                ),
                "utc_wall": "Stored value is UTC wall-clock; *_utc is the raw value.",
                "display_fields": (
                    "Dashboard endpoints also return display_start_time / "
                    "display_end_time, which reproduce the portal's on-screen "
                    "value (raw + 05:30). For IST_WALL columns that is 5h30m "
                    "ahead of reality - a known portal bug, preserved so the "
                    "screens match. Use *_utc for anything real."
                ),
            },
        }
    )


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------


@bp.post("/auth/token")
def issue_token():
    """
    Mint a token using the bootstrap secret.

    Only for when you cannot run ``python -m core.cli mint-token`` (e.g. on
    Vercel).  Requires ``X-Bootstrap-Secret`` to match ``API_BOOTSTRAP_SECRET``;
    if that env var is unset the endpoint is disabled entirely.
    """
    import hmac

    expected = os.getenv("API_BOOTSTRAP_SECRET", "").strip()
    if not expected:
        raise NotFound(
            message=(
                "Token bootstrap is disabled. Set API_BOOTSTRAP_SECRET to enable it, "
                "or mint offline with: python -m core.cli mint-token"
            )
        )

    supplied = request.headers.get("X-Bootstrap-Secret", "").strip()
    if not supplied or not hmac.compare_digest(supplied, expected):
        raise Forbidden(message="Invalid bootstrap secret.", code="invalid_bootstrap_secret")

    body = request.get_json(silent=True) or {}
    subject = (body.get("sub") or "").strip()
    if not subject:
        raise ValidationError(
            message="'sub' is required.",
            details=[{"field": "sub", "issue": "required"}],
        )

    scopes = body.get("scopes") or ["*"]
    if not isinstance(scopes, list):
        raise ValidationError(
            message="'scopes' must be a list.",
            details=[{"field": "scopes", "issue": "must be a list"}],
        )

    unknown = [s for s in scopes if s != "*" and not s.endswith(":*") and s not in ALL_SCOPES]
    if unknown:
        raise ValidationError(
            message="Unknown scope(s).",
            details=[{"field": "scopes", "issue": "unknown", "value": unknown}],
        )

    token, record = mint_token(
        subject=subject,
        scopes=scopes,
        name=body.get("name") or subject,
        note=body.get("note"),
        created_by="bootstrap",
    )
    return ok(
        {
            "token": token,
            "sub": subject,
            "scopes": scopes,
            "jti": record.jti,
            "created_at": record.created_at.isoformat(),
            "warning": "This token is shown once. Store it now.",
        },
        status=201,
    )


@bp.get("/auth/whoami")
@require_auth()
def whoami():
    return ok(current_client().to_dict())


@bp.post("/auth/verify")
@require_auth()
def verify_token():
    body = request.get_json(silent=True) or {}
    raw = body.get("token")
    if not raw:
        raise ValidationError(
            message="'token' is required.",
            details=[{"field": "token", "issue": "required"}],
        )
    client = decode_token(raw)
    return ok({"valid": True, "claims": client.to_dict()})


@bp.get("/auth/scopes")
@require_auth()
def list_scopes():
    return ok(
        {
            "all": list(ALL_SCOPES),
            "read_only_bundle": list(READ_ONLY_SCOPES),
            "wildcards": ["*", "<resource>:*"],
            "granted": list(current_client().scopes),
        }
    )


# ---------------------------------------------------------------------------
# Vocabularies
# ---------------------------------------------------------------------------


@bp.get("/meta/weeks")
@require_auth()
def weeks_endpoint():
    return ok([week_detail(n) for n in week_list()])


@bp.get("/meta/weeks/current")
@require_auth()
def current_week():
    return ok(week_detail(current_week_number()))


@bp.get("/meta/teams")
@require_auth()
def teams():
    return ok(
        [
            {"key": key, "display": team_display(key), "states": list(meta["states"])}
            for key, meta in TEAMS.items()
        ]
    )


@bp.get("/meta/states")
@require_auth()
def states():
    """
    Distinct states.

    ``?source=assignments`` (default) matches the admin/roadvision dropdowns;
    ``surveys`` and ``users`` are also available, and ``known`` returns the list
    hardcoded in templates/admin/reports.html.
    """
    from extensions import db
    from models.db_models import Survey, SurveyAssignment, User

    source = (request.args.get("source") or "assignments").lower()
    if source == "known":
        return ok(list(KNOWN_STATES))

    column = {
        "assignments": SurveyAssignment.state,
        "surveys": Survey.state,
        "users": User.state,
    }.get(source)
    if column is None:
        raise ValidationError(
            message="Unknown source.",
            details=[
                {
                    "field": "source",
                    "issue": "must be one of assignments, surveys, users, known",
                    "value": source,
                }
            ],
        )

    rows = db.session.query(column).distinct().order_by(column).all()
    return ok([row[0] for row in rows if row[0]])


@bp.get("/meta/cycles")
@require_auth()
def cycles():
    from extensions import db
    from models.db_models import Survey

    rows = (
        db.session.query(Survey.cycle_no)
        .distinct()
        .order_by(Survey.cycle_no)
        .all()
    )
    return ok([row[0] for row in rows if row[0] is not None])


@bp.get("/meta/statuses")
@require_auth()
def statuses():
    return ok(
        {
            "survey": [
                {"value": value, "label": SURVEY_STATUS_LABELS[value], "rank": index + 1}
                for index, value in enumerate(SURVEY_STATUSES)
            ],
            "assignment": [{"value": value} for value in ASSIGNMENT_STATUSES],
        }
    )


@bp.get("/meta/roles")
@require_auth()
def roles():
    return ok(list(ROLES))


@bp.get("/meta/days")
@require_auth()
def days():
    return ok(
        {
            "scheduled": [
                {"name": day, "order": WEEKDAY_ORDER[day]} for day in SCHEDULE_DAYS
            ],
            "all": list(ALL_DAYS),
        }
    )


@bp.get("/meta/survey-types")
@require_auth()
def survey_types():
    return ok(list(ALL_SURVEY_TYPES))


@bp.get("/meta/checklists")
@require_auth()
def checklists_endpoint():
    """
    The 20-item pre-survey checklist and the 7-item recording checklist.

    Both are hardcoded HTML in the portal templates and never stored, so this is
    the only way to obtain them programmatically.
    """
    kind = request.args.get("kind")
    if kind:
        items = checklist_data.get_checklist(kind)
        if items is None:
            raise ValidationError(
                message="Unknown checklist kind.",
                details=[
                    {
                        "field": "kind",
                        "issue": f"must be one of {list(checklist_data.CHECKLIST_KINDS)}",
                        "value": kind,
                    }
                ],
            )
        return ok({kind: items})
    return ok(checklist_data.all_checklists())


@bp.get("/meta/defect-fields")
@require_auth()
def defect_fields():
    return ok(
        [
            {"field": field, "side": side, "label": label}
            for field, (side, label) in DEFECT_FIELD_META.items()
        ]
    )


@bp.get("/meta/tasks")
@require_auth()
def tasks():
    """
    Team-leader task keys.

    Note the column names and UI labels disagree - ``task1_completed`` is shown
    as "Raw Video" and ``task2_completed`` as "Final Report".
    """
    return ok(
        [
            {"key": key, "field": field, "timestamp_field": at_field, "label": label}
            for key, (field, at_field, label) in TASK_FIELDS.items()
        ]
    )


@bp.get("/meta/email-types")
@require_auth()
def email_types():
    return ok(list(EMAIL_TYPES))


@bp.get("/meta/filters")
@require_auth("dashboards:read")
def filters():
    """
    The exact dropdown payload a dashboard page builds.

    ``?scope=admin|regional|teamleader|roadvision``.  The regional scope needs
    ``?manager_email=`` because its state and captain lists are scoped to the
    manager's assigned states.
    """
    from extensions import db
    from models.db_models import RegionalManagerState, Survey, SurveyAssignment, User

    scope = (request.args.get("scope") or "admin").lower()
    if scope not in ("admin", "regional", "teamleader", "roadvision"):
        raise ValidationError(
            message="Unknown scope.",
            details=[
                {
                    "field": "scope",
                    "issue": "must be one of admin, regional, teamleader, roadvision",
                    "value": scope,
                }
            ],
        )

    captain_query = User.query.filter_by(role="captain")
    state_values = None

    if scope == "regional":
        manager_email = request.args.get("manager_email")
        if not manager_email:
            raise BadRequest(
                message="scope=regional requires ?manager_email=",
                code="manager_email_required",
            )
        rows = RegionalManagerState.query.filter_by(manager_email=manager_email).all()
        state_values = [row.state for row in rows]
        # regional.py:59 scopes captains by User.region, not User.state.
        captain_query = captain_query.filter(User.region.in_(state_values or [""]))

    if state_values is None:
        rows = (
            db.session.query(SurveyAssignment.state)
            .distinct()
            .order_by(SurveyAssignment.state)
            .all()
        )
        state_values = [row[0] for row in rows if row[0]]

    captains = captain_query.order_by(User.name).all()
    cycle_rows = db.session.query(Survey.cycle_no).distinct().order_by(Survey.cycle_no).all()

    return ok(
        {
            "scope": scope,
            "states": state_values,
            "captains": [
                {"id": c.id, "name": c.name, "email": c.email, "region": c.region}
                for c in captains
            ],
            "cycles": [row[0] for row in cycle_rows if row[0] is not None],
            "weeks": week_list(),
            "statuses": [
                {"value": value, "label": SURVEY_STATUS_LABELS[value]}
                for value in SURVEY_STATUSES
            ],
            "teams": [
                {"key": key, "display": team_display(key)} for key in TEAMS
            ],
        }
    )

# ==========================================================================
# res_users
# /api/v1/users
#
# ``password_hash`` can never be read - it is on an unconditional denylist in
# core/serializers.py, so no combination of ``?fields=`` will surface it.  Passwords
# are set through ``POST /users`` (``password``) or the dedicated password endpoint,
# never by writing the hash directly.
# ==========================================================================

from flask import Blueprint, request

from core.engine import require_auth
from core.config import ROLES
from core.engine import (
    apply_body,
    base_query,
    commit,
    delete_resource,
    get_or_404,
    json_body,
    list_resource,
    restore_resource,
    show_resource,
    update_resource,
    wants_deleted,
)
from core.engine import created, no_content, ok
from core.engine import NotFound, ValidationError
from core.engine import ResourceConfig
from core.engine import Options, spec


MIN_PASSWORD_LENGTH = 8


def user_config():
    from models.db_models import User

    return ResourceConfig(
        model=User,
        table="users",
        spec_name="user",
        searchable=("name", "email", "username"),
        sortable=("id", "name", "email", "username", "role", "region", "state"),
        default_sort="name",
    )


def _validate_role(role):
    if role and role not in ROLES:
        raise ValidationError(
            message=f"Unknown role '{role}'.",
            details=[{"field": "role", "issue": "unknown", "value": role, "allowed": list(ROLES)}],
        )


def _validate_password(password):
    if not password or len(str(password)) < MIN_PASSWORD_LENGTH:
        raise ValidationError(
            message=f"Password must be at least {MIN_PASSWORD_LENGTH} characters.",
            details=[
                {"field": "password", "issue": f"minimum length {MIN_PASSWORD_LENGTH}"}
            ],
        )


@bp.get("/users")
@require_auth("users:read")
def list_users():
    return list_resource(user_config())


@bp.post("/users")
@require_auth("users:write")
def create_user():
    from werkzeug.security import generate_password_hash

    from extensions import db
    from models.db_models import User

    body = json_body()
    password = body.pop("password", None)

    for field in ("name", "email", "role"):
        if not body.get(field):
            raise ValidationError(
                message=f"'{field}' is required.",
                details=[{"field": field, "issue": "required"}],
            )
    _validate_role(body.get("role"))
    _validate_password(password)

    row = User()
    apply_body(row, body, user_config(), partial=True)
    row.password_hash = generate_password_hash(password)

    db.session.add(row)
    commit("create")

    return created(
        spec("user").serialize(row, Options.from_request()),
        location=f"/api/v1/users/{row.id}",
    )


@bp.get("/users/by-email/<path:email>")
@require_auth("users:read")
def user_by_email(email):
    from models.db_models import User

    row = base_query(User, wants_deleted()).filter(User.email.ilike(email)).first()
    if not row:
        raise NotFound(message=f"No user with email '{email}'.")
    return ok(spec("user").serialize(row, Options.from_request()))


@bp.get("/users/<int:user_id>")
@require_auth("users:read")
def show_user(user_id):
    return show_resource(user_config(), user_id)


@bp.patch("/users/<int:user_id>")
@require_auth("users:write")
def patch_user(user_id):
    body = request.get_json(silent=True) or {}
    _validate_role(body.get("role"))
    return update_resource(user_config(), user_id, partial=True)


@bp.put("/users/<int:user_id>")
@require_auth("users:write")
def put_user(user_id):
    body = request.get_json(silent=True) or {}
    _validate_role(body.get("role"))
    return update_resource(user_config(), user_id, partial=False)


@bp.delete("/users/<int:user_id>")
@require_auth("users:write")
def remove_user(user_id):
    # Hard delete is refused when surveys or assignments still reference this
    # user's email - there are no foreign keys to catch it otherwise.
    return delete_resource(user_config(), user_id)


@bp.post("/users/<int:user_id>/restore")
@require_auth("users:write")
def restore_user(user_id):
    return restore_resource(user_config(), user_id)


@bp.post("/users/<int:user_id>/password")
@require_auth("users:write")
def set_password(user_id):
    """
    Set a user's password.

    ``require_current: true`` makes this behave like the portal's own
    change-password screens (routes/captain.py:571, regional.py:536), which verify
    the existing password first.
    """
    from werkzeug.security import check_password_hash, generate_password_hash

    from models.db_models import User

    user = get_or_404(User, user_id, wants_deleted())
    body = json_body()

    new_password = body.get("new_password")
    _validate_password(new_password)

    if body.get("require_current"):
        current = body.get("current_password")
        if not current or not check_password_hash(user.password_hash, current):
            raise ValidationError(
                message="Current password is incorrect.",
                details=[{"field": "current_password", "issue": "incorrect"}],
            )

    confirm = body.get("confirm_password")
    if confirm is not None and confirm != new_password:
        raise ValidationError(
            message="Passwords do not match.",
            details=[{"field": "confirm_password", "issue": "must match new_password"}],
        )

    user.password_hash = generate_password_hash(new_password)
    commit("password")
    return no_content()


@bp.get("/users/<int:user_id>/surveys")
@require_auth("users:read", "surveys:read")
def user_surveys(user_id):
    from models.db_models import Survey, User

    from core.endpoints import survey_config

    user = get_or_404(User, user_id, wants_deleted())
    query = base_query(Survey, wants_deleted()).filter(Survey.captain_email == user.email)
    return list_resource(survey_config(), query=query)


@bp.get("/users/<int:user_id>/assignments")
@require_auth("users:read", "assignments:read")
def user_assignments(user_id):
    from models.db_models import SurveyAssignment, User

    from core.endpoints import assignment_config

    user = get_or_404(User, user_id, wants_deleted())
    query = base_query(SurveyAssignment, wants_deleted()).filter(
        SurveyAssignment.captain_email == user.email
    )
    return list_resource(assignment_config(), query=query)


@bp.get("/users/<int:user_id>/logins")
@require_auth("users:read", "audit:read")
def user_logins(user_id):
    """Login history for one user - see /api/v1/audit/logins for the full feed."""
    from models.db_models import User

    from core.endpoints import login_events_for_user

    user = get_or_404(User, user_id, wants_deleted())
    return login_events_for_user(user)

# ==========================================================================
# res_surveys
# /api/v1/surveys - the central resource.
#
# CRUD plus the derived values the portal computes at render time, aggregate
# statistics, and a reconstructed lifecycle timeline.
#
# ``status`` is deliberately NOT settable through PATCH/PUT.  It only moves via the
# action endpoints in core/res_actions.py, each of which owns its side effects
# (end_time, video_pending_start_time, the mirrored assignment status).  Letting a
# client set it directly would corrupt the upload-duration arithmetic that four
# dashboards depend on.
# ==========================================================================

from flask import Blueprint, request

from core.engine import require_auth
from core.config import DEFECT_COUNT_FIELDS, SURVEY_STATUSES
from core.engine import (
    base_query,
    create_resource,
    delete_resource,
    get_or_404,
    list_resource,
    restore_resource,
    show_resource,
    update_resource,
    wants_deleted,
)
from core.config import (
    admin_rank_case,
    attach_derived,
    regional_rank_case,
    roadvision_rank_case,
    teamleader_rank_case,
)
from core.engine import ok
from core.engine import ValidationError
from core.engine import ResourceConfig, parse_int, survey_custom_filters
from core.engine import Options, spec, survey_derived
from core.config import iso, as_utc, semantic_for


#: status is excluded from direct writes; see the module docstring.
PROTECTED_SURVEY_FIELDS = ("status",)


def survey_config():
    from models.db_models import Survey

    return ResourceConfig(
        model=Survey,
        table="surveys",
        spec_name="survey",
        searchable=("stretch_code", "section_no", "upc_code", "captain_name"),
        sortable=(
            "id",
            "start_time",
            "end_time",
            "section_no",
            "upc_code",
            "cycle_no",
            "state",
            "captain_name",
            "captain_email",
            "status",
            "section_length",
            "video_upload_time",
            "survey_day",
            "roadvision_completed",
        ),
        sort_presets={
            "admin_rank": lambda: [admin_rank_case(), _start_desc()],
            "regional_rank": lambda: [regional_rank_case(), _start_desc()],
            "teamleader_rank": lambda: [teamleader_rank_case(), _start_desc()],
            "roadvision_rank": lambda: [roadvision_rank_case(), _start_desc()],
        },
        custom_filters=survey_custom_filters(),
        default_sort="-start_time",
    )


def _start_desc():
    from models.db_models import Survey

    return Survey.start_time.desc()


# ---------------------------------------------------------------------------
# Collection
# ---------------------------------------------------------------------------


@bp.get("/surveys")
@require_auth("surveys:read")
def list_surveys():
    """
    List surveys.

    Supports the full filter set, the four dashboard sort presets, and
    ``?derived=true`` to attach the computed block to each row.
    """
    derived = request.args.get("derived", "").lower() in ("1", "true", "yes")

    def transform(rows):
        if derived:
            attach_derived(rows)

    return list_resource(survey_config(), transform=transform if derived else None)


@bp.post("/surveys")
@require_auth("surveys:write")
def create_survey():
    """
    Create a survey row directly.

    This is the raw escape hatch.  For the real workflow use
    ``POST /surveys/start``, which applies the weekly-duplicate guard, derives
    cycle_no and updates the assignment.
    """
    from core.config import SURVEY_ONGOING

    return create_resource(
        survey_config(),
        location_prefix="/api/v1/surveys",
        required=("captain_email", "section_no"),
        forbidden=PROTECTED_SURVEY_FIELDS,
        defaults={"status": SURVEY_ONGOING, "cycle_no": 1},
    )


@bp.get("/surveys/count")
@require_auth("surveys:read")
def count_surveys():
    from core.engine import ListParams

    cfg = survey_config()
    params = ListParams.from_request(cfg)
    query = params.apply_filters(base_query(cfg.model, wants_deleted()))
    return ok({"count": query.count(), "filters_applied": params.filters_applied})


@bp.get("/surveys/stats")
@require_auth("surveys:read")
def survey_stats():
    """
    Aggregate over any filtered set.

    ``?group_by=state,status`` and ``?metric=count|km|hours`` (comma-separated).
    This is what makes the other site able to build its own charts without
    pulling every row.
    """
    from extensions import db
    from models.db_models import Survey

    from core.engine import ListParams

    allowed_groups = {
        "status": Survey.status,
        "state": Survey.state,
        "cycle_no": Survey.cycle_no,
        "captain_email": Survey.captain_email,
        "captain_name": Survey.captain_name,
        "survey_day": Survey.survey_day,
        "survey_type": Survey.survey_type,
        "section_no": Survey.section_no,
        "upc_code": Survey.upc_code,
        "nh_number": Survey.nh_number,
        "ro": Survey.ro,
        "piu": Survey.piu,
    }

    raw_groups = request.args.get("group_by", "status")
    group_names = [g.strip() for g in raw_groups.split(",") if g.strip()]
    unknown = [g for g in group_names if g not in allowed_groups and g not in ("team", "week")]
    if unknown:
        raise ValidationError(
            message="Unknown group_by field(s).",
            details=[
                {
                    "field": "group_by",
                    "issue": "not groupable",
                    "value": unknown,
                    "allowed": sorted(list(allowed_groups) + ["team", "week"]),
                }
            ],
        )

    cfg = survey_config()
    params = ListParams.from_request(cfg, extra_allowed={"group_by", "metric"})
    query = params.apply_filters(base_query(cfg.model, wants_deleted()))

    # "team" and "week" have no column, so those two group in Python.
    python_groups = [g for g in group_names if g in ("team", "week")]
    sql_groups = [g for g in group_names if g not in ("team", "week")]

    if python_groups:
        return _stats_in_python(query.all(), group_names, params)

    columns = [allowed_groups[g] for g in sql_groups]
    rows = (
        query.with_entities(
            *columns,
            db.func.count(Survey.id),
            db.func.sum(db.func.coalesce(Survey.section_length, 0)),
        )
        .group_by(*columns)
        .all()
    )

    groups = []
    for row in rows:
        key = dict(zip(sql_groups, row[: len(sql_groups)]))
        groups.append(
            {
                "key": key,
                "count": int(row[-2] or 0),
                "km": round(float(row[-1] or 0), 2),
            }
        )
    groups.sort(key=lambda g: -g["count"])

    return ok(
        {
            "group_by": group_names,
            "groups": groups,
            "totals": {
                "count": sum(g["count"] for g in groups),
                "km": round(sum(g["km"] for g in groups), 2),
            },
        },
        meta=params.meta(),
    )


def _stats_in_python(rows, group_names, params):
    from core.config import team_for_state
    from core.config import survey_duration_minutes
    from core.config import week_of

    buckets = {}
    for survey in rows:
        key = {}
        for name in group_names:
            if name == "team":
                key["team"] = team_for_state(survey.state)
            elif name == "week":
                key["week"] = week_of(survey.start_time)
            else:
                key[name] = getattr(survey, name, None)
        signature = tuple(sorted(key.items(), key=lambda kv: str(kv[0])))
        bucket = buckets.setdefault(
            signature, {"key": key, "count": 0, "km": 0.0, "minutes": 0.0}
        )
        bucket["count"] += 1
        bucket["km"] += survey.section_length or 0
        minutes = survey_duration_minutes(survey)
        if minutes:
            bucket["minutes"] += minutes

    groups = [
        {
            "key": b["key"],
            "count": b["count"],
            "km": round(b["km"], 2),
            "hours": round(b["minutes"] / 60, 2),
        }
        for b in buckets.values()
    ]
    groups.sort(key=lambda g: -g["count"])

    return ok(
        {
            "group_by": group_names,
            "groups": groups,
            "totals": {
                "count": sum(g["count"] for g in groups),
                "km": round(sum(g["km"] for g in groups), 2),
                "hours": round(sum(g["hours"] for g in groups), 2),
            },
        },
        meta=params.meta(),
    )


# ---------------------------------------------------------------------------
# Item
# ---------------------------------------------------------------------------


@bp.get("/surveys/<int:survey_id>")
@require_auth("surveys:read")
def show_survey(survey_id):
    def extra(row):
        if request.args.get("derived", "").lower() in ("1", "true", "yes"):
            return {"derived": survey_derived(row)}
        return None

    return show_resource(survey_config(), survey_id, extra=extra)


@bp.patch("/surveys/<int:survey_id>")
@require_auth("surveys:write")
def patch_survey(survey_id):
    return update_resource(
        survey_config(), survey_id, partial=True, forbidden=PROTECTED_SURVEY_FIELDS
    )


@bp.put("/surveys/<int:survey_id>")
@require_auth("surveys:write")
def put_survey(survey_id):
    return update_resource(
        survey_config(), survey_id, partial=False, forbidden=PROTECTED_SURVEY_FIELDS
    )


@bp.delete("/surveys/<int:survey_id>")
@require_auth("surveys:write")
def remove_survey(survey_id):
    return delete_resource(survey_config(), survey_id)


@bp.post("/surveys/<int:survey_id>/restore")
@require_auth("surveys:write")
def restore_survey(survey_id):
    return restore_resource(survey_config(), survey_id)


@bp.get("/surveys/<int:survey_id>/derived")
@require_auth("surveys:read")
def survey_derived_endpoint(survey_id):
    from models.db_models import Survey

    survey = get_or_404(Survey, survey_id, wants_deleted())
    from core.config import team_display, team_for_state
    from core.config import scheduled_day_map, survey_ref_id
    from core.config import week_of

    data = survey_derived(survey)
    team = team_for_state(survey.state)
    data.update(
        {
            "scheduled_day": scheduled_day_map([survey.section_no]).get(survey.section_no)
            or survey.survey_day,
            "week_no": week_of(survey.start_time),
            "team": team,
            "team_display": team_display(team) if team else None,
            "survey_ref_id": survey_ref_id(survey) if survey.upc_code else None,
        }
    )
    return ok(data)


@bp.get("/surveys/<int:survey_id>/timeline")
@require_auth("surveys:read")
def survey_timeline(survey_id):
    """
    Lifecycle reconstructed from the timestamp columns.

    The portal stores no event history, so this is inferred - each step appears
    only if its timestamp is populated.
    """
    from models.db_models import Survey

    survey = get_or_404(Survey, survey_id, wants_deleted())

    def point(label, value, column, note=None):
        if value is None:
            return None
        semantic = semantic_for("surveys", column)
        return {
            "step": label,
            "at_utc": iso(as_utc(value, semantic)),
            "at_raw": value.isoformat(),
            "note": note,
        }

    steps = [
        point("started", survey.start_time, "start_time"),
        point("survey_ended", survey.end_time, "end_time", "PDF uploaded, video pending"),
        point(
            "video_pending_since",
            survey.video_pending_start_time,
            "video_pending_start_time",
        ),
        point("video_uploaded", survey.video_upload_time, "video_upload_time"),
        point(
            "survey_form_completed",
            survey.survey_form_completed_at,
            "survey_form_completed_at",
            "team leader",
        ),
        point("raw_video_done", survey.task1_completed_at, "task1_completed_at", "task1"),
        point("final_report_done", survey.task2_completed_at, "task2_completed_at", "task2"),
        point(
            "roadvision_reviewed",
            survey.roadvision_completed_at,
            "roadvision_completed_at",
        ),
        point(
            "video_count_checked",
            getattr(survey, "video_count_checked_at", None),
            "video_count_checked_at",
            "automated bucket comparison",
        ),
    ]
    steps = [step for step in steps if step]
    steps.sort(key=lambda s: s["at_utc"] or "")

    return ok(
        {
            "survey_id": survey.id,
            "status": survey.status,
            "steps": steps,
        }
    )


@bp.get("/surveys/<int:survey_id>/media")
@require_auth("surveys:read")
def survey_media(survey_id):
    from models.db_models import Survey

    survey = get_or_404(Survey, survey_id, wants_deleted())
    return ok(
        {
            "survey_id": survey.id,
            "dashcam_photo": survey.dashcam_photo,
            "end_survey_pdf": survey.end_survey_pdf,
            # Read by the regional/teamleader detail templates but never written
            # by any route - it is effectively always empty.
            "end_survey_photo": survey.end_survey_photo,
            "settings_photo": survey.settings_photo,
            "pdf_reupload": {
                "required": bool(survey.pdf_reupload_required),
                "reason": survey.pdf_reupload_reason,
                "count": survey.pdf_reupload_count or 0,
            },
        }
    )


@bp.get("/surveys/<int:survey_id>/defect-counts")
@require_auth("surveys:read")
def survey_defect_counts(survey_id):
    from models.db_models import Survey

    from core.config import defect_counts

    survey = get_or_404(Survey, survey_id, wants_deleted())
    payload = defect_counts(survey)
    payload["total"] = sum(getattr(survey, field, 0) or 0 for field in DEFECT_COUNT_FIELDS)
    payload["video_uploaded"] = bool(survey.video_uploaded)
    payload["video_count_matched"] = getattr(survey, "video_count_matched", None)
    return ok(payload)


@bp.get("/surveys/statuses")
@require_auth("surveys:read")
def survey_status_counts():
    """Row count per status over the filtered set - the dashboard KPI numbers."""
    from extensions import db
    from models.db_models import Survey

    from core.engine import ListParams

    cfg = survey_config()
    params = ListParams.from_request(cfg)
    query = params.apply_filters(base_query(cfg.model, wants_deleted()))
    rows = query.with_entities(Survey.status, db.func.count(Survey.id)).group_by(
        Survey.status
    ).all()
    counts = {status: 0 for status in SURVEY_STATUSES}
    counts.update({status: int(total) for status, total in rows if status})
    counts["total"] = sum(int(total) for _status, total in rows)
    return ok(counts)


_ = (Options, spec, parse_int)

# ==========================================================================
# res_actions
# Workflow action endpoints.
#
# Everything that moves a survey through its lifecycle.  Separate from
# core/res_surveys.py because these are operations, not CRUD: each owns side
# effects, validates the transition, and returns 409 rather than silently doing the
# wrong thing.
#
# ``POST /surveys/start`` is the only genuinely non-idempotent call here - a retry
# after a network timeout would create a second survey and increment ``cycle_no``
# again - so it honours an ``Idempotency-Key`` header.  The rest are naturally
# idempotent once the transition validator is in place: calling
# ``groundwork-complete`` twice gives 200 then 409, and the end state is the same
# either way.
# ==========================================================================

import hashlib
import json

from flask import Blueprint, request

from core import engine as svc_survey
from core.engine import current_client, require_auth, resolve_actor
from core.engine import get_or_404, json_body, wants_deleted
from core.engine import created, ok
from core.engine import Conflict, NotFound, ValidationError
from core.engine import Options, spec



def _survey_json(survey):
    return spec("survey").serialize(survey, Options.from_request())


# ---------------------------------------------------------------------------
# Idempotency
# ---------------------------------------------------------------------------


def _idempotency_guard(endpoint, body):
    """
    Returns ``(replayed_response | None, claim_row | None)``.

    DB-backed because serverless instances share no memory - two concurrent
    invocations must not both win the claim.
    """
    key = request.headers.get("Idempotency-Key")
    if not key:
        return None, None

    from extensions import db

    from core.models import IdempotencyKey

    digest = hashlib.sha256(
        json.dumps(body, sort_keys=True, default=str).encode()
    ).hexdigest()

    existing = IdempotencyKey.query.filter_by(key=key, endpoint=endpoint).first()
    if existing:
        if existing.request_hash != digest:
            raise ValidationError(
                message=(
                    "This Idempotency-Key was already used with a different body."
                ),
                code="idempotency_conflict",
                details=[{"field": "Idempotency-Key", "issue": "reused with new payload"}],
            )
        if existing.response is not None:
            return (
                ok(existing.response, headers={"Idempotency-Replayed": "true"}),
                None,
            )
        raise Conflict(
            message="An identical request is still in flight.",
            code="idempotency_in_progress",
        )

    claim = IdempotencyKey(key=key, endpoint=endpoint, request_hash=digest)
    db.session.add(claim)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise Conflict(
            message="An identical request is already being processed.",
            code="idempotency_in_progress",
        )
    return None, claim


def _record_idempotent_result(claim, payload, status=200):
    if not claim:
        return
    from extensions import db

    try:
        claim.response = payload
        claim.status_code = status
        db.session.commit()
    except Exception:
        db.session.rollback()


# ---------------------------------------------------------------------------
# Start
# ---------------------------------------------------------------------------


@bp.post("/surveys/start")
@require_auth("actions:write")
def start_survey():
    """
    Start a survey - the API equivalent of a captain submitting the checklist.

    Body: ``assignment_id`` (required), ``survey_type``, ``survey_day``,
    ``captain_email`` (or a user-bound token), ``dashcam_photo``, ``is_resurvey``,
    ``force``.

    409 ``survey_already_started`` if this section was already surveyed this week
    and no re-survey has been approved.
    """
    from models.db_models import SurveyAssignment

    body = json_body()
    replay, claim = _idempotency_guard("POST /surveys/start", body)
    if replay:
        return replay

    assignment_id = body.get("assignment_id")
    if not assignment_id:
        raise ValidationError(
            message="'assignment_id' is required.",
            details=[{"field": "assignment_id", "issue": "required"}],
        )

    assignment = SurveyAssignment.query.get(assignment_id)
    if not assignment:
        raise NotFound(message=f"Assignment {assignment_id} not found.")

    captain = resolve_actor(
        explicit_email=body.get("captain_email") or assignment.captain_email,
        roles=("captain", "backup_captain", "admin"),
    )

    survey_type = body.get("survey_type") or assignment.survey_type or "Day"

    active = svc_survey.active_survey_for(captain.email)
    if active and not body.get("force"):
        raise Conflict(
            message=(
                f"{captain.name} already has survey {active.id} in progress "
                f"({active.status}). Finish it first."
            ),
            code="captain_already_surveying",
            details=[{"field": "captain_email", "issue": "has an active survey",
                      "value": active.id}],
        )

    survey = svc_survey.start_survey(
        assignment=assignment,
        captain=captain,
        survey_type=survey_type,
        survey_day=body.get("survey_day"),
        dashcam_photo=body.get("dashcam_photo"),
        is_resurvey=bool(body.get("is_resurvey")),
        actor_role=captain.role,
        force=bool(body.get("force")),
    )

    payload = _survey_json(survey)
    _record_idempotent_result(claim, payload, 201)
    return created(payload, location=f"/api/v1/surveys/{survey.id}")


# ---------------------------------------------------------------------------
# Lifecycle
# ---------------------------------------------------------------------------


@bp.post("/surveys/<int:survey_id>/groundwork-complete")
@require_auth("actions:write")
def groundwork_complete(survey_id):
    return ok(_survey_json(svc_survey.groundwork_complete(survey_id)))


@bp.post("/surveys/<int:survey_id>/complete")
@require_auth("actions:write")
def complete(survey_id):
    """
    Finish field work.

    Either supply ``pdf_url`` in a JSON body, or upload the file itself as
    multipart ``survey_pdf`` and it will be optimised and pushed to Drive first.
    """
    from models.db_models import Survey

    survey = get_or_404(Survey, survey_id, wants_deleted())

    pdf_url = None
    actor_role = None

    if request.files.get("survey_pdf"):
        from core.engine import upload_survey_pdf

        pdf_url = upload_survey_pdf(request.files["survey_pdf"], survey)["view_url"]
        actor_role = request.form.get("actor_role")
    else:
        body = json_body()
        pdf_url = body.get("pdf_url") or body.get("end_survey_pdf")
        actor_role = body.get("actor_role")

    return ok(_survey_json(svc_survey.complete_survey(survey_id, pdf_url, actor_role)))


@bp.post("/surveys/<int:survey_id>/video-counts")
@require_auth("actions:write")
def video_counts(survey_id):
    """Record the 8 LHS/RHS defect counts and mark the survey completed."""
    return ok(_survey_json(svc_survey.submit_video_counts(survey_id, json_body())))


@bp.post("/surveys/<int:survey_id>/cancel")
@require_auth("actions:write")
def cancel(survey_id):
    body = json_body(required=False)
    return ok(_survey_json(svc_survey.cancel_survey(survey_id, body.get("reason"))))


# ---------------------------------------------------------------------------
# Re-survey
# ---------------------------------------------------------------------------


@bp.post("/surveys/<int:survey_id>/resurvey/request")
@require_auth("actions:write")
def resurvey_request(survey_id):
    return ok(_survey_json(svc_survey.request_resurvey(survey_id)))


@bp.post("/surveys/<int:survey_id>/resurvey/approve")
@require_auth("actions:write")
def resurvey_approve(survey_id):
    return ok(_survey_json(svc_survey.approve_resurvey(survey_id)))


@bp.post("/surveys/<int:survey_id>/resurvey/reject")
@require_auth("actions:write")
def resurvey_reject(survey_id):
    return ok(_survey_json(svc_survey.reject_resurvey(survey_id)))


@bp.get("/surveys/resurvey/pending")
@require_auth("surveys:read")
def pending_resurveys():
    """The "Re-Survey Requests" scroller on the admin dashboard."""
    from models.db_models import Survey

    from core.engine import base_query

    rows = (
        base_query(Survey)
        .filter(Survey.resurvey_requested.is_(True), Survey.resurvey_approved.isnot(True))
        .order_by(Survey.id.desc())
        .all()
    )
    options = Options.from_request()
    return ok([spec("survey").serialize(row, options) for row in rows])


# ---------------------------------------------------------------------------
# PDF re-upload
# ---------------------------------------------------------------------------


@bp.post("/surveys/<int:survey_id>/pdf-reupload/request")
@require_auth("actions:write")
def pdf_reupload_request(survey_id):
    body = json_body()
    return ok(
        _survey_json(svc_survey.request_pdf_reupload(survey_id, body.get("reason")))
    )


@bp.post("/surveys/<int:survey_id>/pdf-reupload")
@require_auth("actions:write")
def pdf_reupload(survey_id):
    from models.db_models import Survey

    survey = get_or_404(Survey, survey_id, wants_deleted())

    if request.files.get("survey_pdf"):
        from core.engine import upload_survey_pdf

        pdf_url = upload_survey_pdf(request.files["survey_pdf"], survey)["view_url"]
    else:
        pdf_url = json_body().get("pdf_url")

    if not pdf_url:
        raise ValidationError(
            message="Provide 'pdf_url' or upload a multipart 'survey_pdf' file.",
            details=[{"field": "pdf_url", "issue": "required"}],
        )

    return ok(_survey_json(svc_survey.complete_pdf_reupload(survey_id, pdf_url)))


# ---------------------------------------------------------------------------
# Review / tasks
# ---------------------------------------------------------------------------


@bp.post("/surveys/<int:survey_id>/roadvision-review")
@require_auth("actions:write")
def roadvision_review(survey_id):
    body = json_body()
    return ok(_survey_json(svc_survey.roadvision_review(survey_id, body.get("remark"))))


@bp.post("/surveys/<int:survey_id>/tasks/<task>")
@require_auth("actions:write")
def toggle_task(survey_id, task):
    """
    Mark a team-leader task done.

    ``task`` is ``task1`` (labelled "Raw Video" in the UI), ``task2``
    ("Final Report") or ``survey_form`` ("Survey Form").
    """
    body = json_body(required=False)
    completed = body.get("completed", True)
    allow_reset = str(request.args.get("allow_reset", "")).lower() in ("1", "true", "yes")
    return ok(
        _survey_json(
            svc_survey.toggle_task(survey_id, task, completed=bool(completed),
                                   allow_reset=allow_reset)
        )
    )


@bp.post("/surveys/<int:survey_id>/visibility")
@require_auth("surveys:write")
def set_visibility(survey_id):
    body = json_body()
    return ok(
        _survey_json(
            svc_survey.set_visibility(
                survey_id,
                show_on_dashboard=body.get("show_on_dashboard"),
                show_in_teamleader_dashboard=body.get("show_in_teamleader_dashboard"),
            )
        )
    )


@bp.get("/surveys/<int:survey_id>/allowed-transitions")
@require_auth("surveys:read")
def allowed_transitions(survey_id):
    """What this survey can do next - useful for building a UI on the other site."""
    from models.db_models import Survey

    from core.config import ALLOWED_STATUS_TRANSITIONS

    survey = get_or_404(Survey, survey_id, wants_deleted())
    allowed = sorted(ALLOWED_STATUS_TRANSITIONS.get(survey.status, set()))

    endpoints = {
        "groundwork_completed": f"POST /api/v1/surveys/{survey.id}/groundwork-complete",
        "video_pending": f"POST /api/v1/surveys/{survey.id}/complete",
        "completed": f"POST /api/v1/surveys/{survey.id}/video-counts",
        "cancelled": f"POST /api/v1/surveys/{survey.id}/cancel",
    }

    return ok(
        {
            "survey_id": survey.id,
            "status": survey.status,
            "allowed_next": allowed,
            "endpoints": {state: endpoints[state] for state in allowed if state in endpoints},
            "available_actions": {
                "request_resurvey": not survey.resurvey_requested,
                "approve_resurvey": bool(survey.resurvey_requested)
                and not survey.resurvey_approved,
                "roadvision_review": not survey.roadvision_completed,
                "request_pdf_reupload": bool(survey.end_survey_pdf),
                "complete_pdf_reupload": bool(survey.pdf_reupload_required),
            },
        }
    )


_ = current_client

# ==========================================================================
# res_assignments
# /api/v1/assignments - the weekly roster (``survey_assignments``).
#
# Also carries the missed-survey bookkeeping the admin dashboard reads:
# ``status``, ``missed_reason``, ``alert_acknowledged``, ``survey_enabled``.
# ==========================================================================

from flask import Blueprint, request

from core.engine import require_auth
from core.config import ASSIGNMENT_STATUSES, SCHEDULE_DAYS
from core.engine import (
    base_query,
    commit,
    create_resource,
    delete_resource,
    get_or_404,
    json_body,
    list_resource,
    restore_resource,
    show_resource,
    update_resource,
    wants_deleted,
)
from core.config import day_counts, day_counts_global, day_order_case
from core.engine import ok
from core.engine import ValidationError
from core.engine import ResourceConfig, assignment_custom_filters
from core.engine import Options, spec



def assignment_config():
    from models.db_models import SurveyAssignment

    return ResourceConfig(
        model=SurveyAssignment,
        table="survey_assignments",
        spec_name="assignment",
        searchable=("stretch_code", "section_no", "upc_code", "main_person"),
        sortable=(
            "id",
            "section_no",
            "stretch_code",
            "state",
            "survey_day",
            "status",
            "captain_email",
            "main_person",
            "section_length",
            "upc_code",
            "survey_enabled",
        ),
        sort_presets={
            "day_order": lambda: [day_order_case(), _state_then_section()],
        },
        custom_filters=assignment_custom_filters(),
        default_sort="section_no",
    )


def _state_then_section():
    from models.db_models import SurveyAssignment

    return SurveyAssignment.state.asc(), SurveyAssignment.section_no.asc()


@bp.get("/assignments")
@require_auth("assignments:read")
def list_assignments():
    return list_resource(assignment_config())


@bp.post("/assignments")
@require_auth("assignments:write")
def create_assignment():
    return create_resource(
        assignment_config(),
        location_prefix="/api/v1/assignments",
        required=("captain_email", "section_no"),
        defaults={"status": "assigned", "survey_enabled": False},
    )


@bp.get("/assignments/schedule-summary")
@require_auth("assignments:read")
def schedule_summary():
    """
    Mon-Fri counts, matching the schedules pages.

    ``?respect_filters=true`` counts the filtered set (what regional and
    teamleader do); the default counts everything, reproducing the admin page,
    whose cards deliberately ignore the active filters (admin.py:707-725).
    """
    from core.engine import ListParams

    cfg = assignment_config()
    params = ListParams.from_request(cfg, extra_allowed={"respect_filters"})
    respect = request.args.get("respect_filters", "false").lower() in ("1", "true", "yes")

    if respect:
        rows = params.apply_filters(base_query(cfg.model, wants_deleted())).all()
        counts = day_counts(rows)
    else:
        counts = day_counts_global()

    counts["days"] = list(SCHEDULE_DAYS)
    counts["respects_filters"] = respect
    return ok(counts, meta=params.meta())


@bp.post("/assignments/bulk")
@require_auth("assignments:write")
def bulk_assignments():
    """
    Upsert many assignments at once, keyed on ``section_no`` or ``id``.

    This is how the other site can push a whole week's roster in one call.
    Reports per-row errors instead of failing the entire batch.
    """
    from extensions import db
    from models.db_models import SurveyAssignment

    from core.engine import apply_body

    body = json_body()
    items = body.get("items")
    if not isinstance(items, list) or not items:
        raise ValidationError(
            message="'items' must be a non-empty list.",
            details=[{"field": "items", "issue": "required, non-empty list"}],
        )
    if len(items) > 1000:
        raise ValidationError(
            message="At most 1000 items per call.",
            details=[{"field": "items", "issue": "max 1000", "value": len(items)}],
        )

    key_field = body.get("key", "section_no")
    if key_field not in ("section_no", "id"):
        raise ValidationError(
            message="'key' must be 'section_no' or 'id'.",
            details=[{"field": "key", "issue": "invalid", "value": key_field}],
        )
    mode = body.get("mode", "upsert")
    if mode not in ("upsert", "insert", "update"):
        raise ValidationError(
            message="'mode' must be upsert, insert or update.",
            details=[{"field": "mode", "issue": "invalid", "value": mode}],
        )

    cfg = assignment_config()
    created_count = updated_count = skipped = 0
    errors = []

    for index, item in enumerate(items):
        if not isinstance(item, dict):
            errors.append({"index": index, "error": "each item must be an object"})
            continue
        key_value = item.get(key_field)
        if key_value in (None, ""):
            errors.append({"index": index, "error": f"missing '{key_field}'"})
            continue

        try:
            existing = (
                SurveyAssignment.query.filter_by(**{key_field: key_value})
                .order_by(SurveyAssignment.id)
                .first()
            )
            if existing and mode == "insert":
                skipped += 1
                continue
            if not existing and mode == "update":
                skipped += 1
                continue

            row = existing or SurveyAssignment()
            payload = {k: v for k, v in item.items() if k != "id"}
            apply_body(row, payload, cfg, partial=True)
            if existing:
                updated_count += 1
            else:
                db.session.add(row)
                created_count += 1
        except Exception as exc:
            db.session.rollback()
            errors.append({"index": index, "error": str(exc)[:300]})

    commit("bulk")
    return ok(
        {
            "created": created_count,
            "updated": updated_count,
            "skipped": skipped,
            "errors": errors,
        }
    )


@bp.post("/assignments/weekly-reset")
@require_auth("assignments:write")
def weekly_reset():
    """
    The Monday reset, as an explicit operation.

    The portal runs this as a side effect of loading ``GET /admin``
    (routes/admin.py:61-84).  Exposing it as a POST means a machine client can
    trigger it deliberately instead of by accident, and ``dry_run`` lets you see
    the blast radius first.
    """
    from core.engine import run_weekly_reset

    body = json_body(required=False)
    result = run_weekly_reset(
        force=bool(body.get("force")), dry_run=bool(body.get("dry_run"))
    )
    return ok(result)


@bp.get("/assignments/<int:assignment_id>")
@require_auth("assignments:read")
def show_assignment(assignment_id):
    return show_resource(assignment_config(), assignment_id)


@bp.patch("/assignments/<int:assignment_id>")
@require_auth("assignments:write")
def patch_assignment(assignment_id):
    _validate_status(request.get_json(silent=True) or {})
    return update_resource(assignment_config(), assignment_id, partial=True)


@bp.put("/assignments/<int:assignment_id>")
@require_auth("assignments:write")
def put_assignment(assignment_id):
    _validate_status(request.get_json(silent=True) or {})
    return update_resource(assignment_config(), assignment_id, partial=False)


@bp.delete("/assignments/<int:assignment_id>")
@require_auth("assignments:write")
def remove_assignment(assignment_id):
    return delete_resource(assignment_config(), assignment_id)


@bp.post("/assignments/<int:assignment_id>/restore")
@require_auth("assignments:write")
def restore_assignment(assignment_id):
    return restore_resource(assignment_config(), assignment_id)


def _validate_status(body):
    status = body.get("status")
    if status and status not in ASSIGNMENT_STATUSES:
        raise ValidationError(
            message=f"Unknown assignment status '{status}'.",
            details=[
                {
                    "field": "status",
                    "issue": "unknown",
                    "value": status,
                    "allowed": list(ASSIGNMENT_STATUSES),
                }
            ],
        )


# ---------------------------------------------------------------------------
# Actions
# ---------------------------------------------------------------------------


@bp.post("/assignments/<int:assignment_id>/unable-to-survey")
@require_auth("actions:write")
def unable_to_survey(assignment_id):
    """Record why a captain could not survey - routes/captain.py:826-840."""
    from models.db_models import SurveyAssignment

    assignment = get_or_404(SurveyAssignment, assignment_id, wants_deleted())
    body = json_body()
    reason = (body.get("reason") or "").strip()
    if not reason:
        raise ValidationError(
            message="'reason' is required.",
            details=[{"field": "reason", "issue": "required, non-empty"}],
        )

    assignment.missed_reason = reason
    commit("unable")
    return ok(spec("assignment").serialize(assignment, Options.from_request()))


@bp.post("/assignments/<int:assignment_id>/acknowledge-alert")
@require_auth("alerts:write")
def acknowledge_alert(assignment_id):
    """
    Acknowledge a missed-survey alert.

    Default scope ``captain_day`` mirrors routes/admin.py:634-641, which
    acknowledges every assignment sharing the captain and day - not just this one.
    """
    from models.db_models import SurveyAssignment

    assignment = get_or_404(SurveyAssignment, assignment_id, wants_deleted())
    body = json_body(required=False)
    scope = body.get("scope", "captain_day")

    if scope == "self":
        targets = [assignment]
    elif scope == "captain_day":
        targets = SurveyAssignment.query.filter_by(
            captain_email=assignment.captain_email, survey_day=assignment.survey_day
        ).all()
    else:
        raise ValidationError(
            message="'scope' must be 'self' or 'captain_day'.",
            details=[{"field": "scope", "issue": "invalid", "value": scope}],
        )

    for target in targets:
        target.alert_acknowledged = True
    commit("acknowledge")

    return ok({"updated_count": len(targets), "scope": scope})


@bp.post("/assignments/<int:assignment_id>/enable")
@require_auth("assignments:write")
def enable_assignment(assignment_id):
    return _set_enabled(assignment_id, True)


@bp.post("/assignments/<int:assignment_id>/disable")
@require_auth("assignments:write")
def disable_assignment(assignment_id):
    return _set_enabled(assignment_id, False)


def _set_enabled(assignment_id, value):
    from models.db_models import SurveyAssignment

    assignment = get_or_404(SurveyAssignment, assignment_id, wants_deleted())
    assignment.survey_enabled = value
    commit("enable")
    return ok(spec("assignment").serialize(assignment, Options.from_request()))

# ==========================================================================
# res_simple
# The three small resources: legacy schedules, equipment, and regional-manager
# state mappings.
#
# ``survey_schedule`` is superseded by ``survey_assignments`` and is only read by
# ``GET /captain`` (routes/captain.py:63-70).  ``equipment`` is never queried by any
# portal route at all - dashcam and powerbank codes are denormalised onto the
# assignment rows - but it is exposed here for completeness.
# ==========================================================================

from flask import Blueprint

from core.engine import require_auth
from core.engine import (
    base_query,
    commit,
    create_resource,
    delete_resource,
    json_body,
    list_resource,
    show_resource,
    update_resource,
    wants_deleted,
)
from core.engine import ok
from core.engine import NotFound, ValidationError
from core.engine import ResourceConfig
from core.engine import Options, spec



# ---------------------------------------------------------------------------
# Schedules (legacy)
# ---------------------------------------------------------------------------


def schedule_config():
    from models.db_models import SurveySchedule

    return ResourceConfig(
        model=SurveySchedule,
        table="survey_schedule",
        spec_name="schedule",
        searchable=("stretch_code", "main_person"),
        sortable=("id", "captain_email", "stretch_code", "state", "survey_day"),
        default_sort="id",
    )


@bp.get("/schedules")
@require_auth("schedules:read")
def list_schedules():
    return list_resource(schedule_config())


@bp.post("/schedules")
@require_auth("schedules:write")
def create_schedule():
    return create_resource(
        schedule_config(), location_prefix="/api/v1/schedules", required=("captain_email",)
    )


@bp.get("/schedules/by-captain/<path:email>")
@require_auth("schedules:read")
def schedule_by_captain(email):
    """The single row GET /captain reads (routes/captain.py:63-65)."""
    from models.db_models import SurveySchedule

    row = (
        base_query(SurveySchedule, wants_deleted())
        .filter(SurveySchedule.captain_email.ilike(email))
        .first()
    )
    if not row:
        raise NotFound(message=f"No schedule for captain '{email}'.")
    return ok(spec("schedule").serialize(row, Options.from_request()))


@bp.get("/schedules/<int:schedule_id>")
@require_auth("schedules:read")
def show_schedule(schedule_id):
    return show_resource(schedule_config(), schedule_id)


@bp.patch("/schedules/<int:schedule_id>")
@require_auth("schedules:write")
def patch_schedule(schedule_id):
    return update_resource(schedule_config(), schedule_id, partial=True)


@bp.put("/schedules/<int:schedule_id>")
@require_auth("schedules:write")
def put_schedule(schedule_id):
    return update_resource(schedule_config(), schedule_id, partial=False)


@bp.delete("/schedules/<int:schedule_id>")
@require_auth("schedules:write")
def remove_schedule(schedule_id):
    return delete_resource(schedule_config(), schedule_id)


# ---------------------------------------------------------------------------
# Equipment
# ---------------------------------------------------------------------------


def equipment_config():
    from models.db_models import Equipment

    return ResourceConfig(
        model=Equipment,
        table="equipment",
        spec_name="equipment",
        searchable=("equipment_code", "equipment_type"),
        sortable=("id", "equipment_code", "equipment_type"),
        default_sort="equipment_code",
    )


@bp.get("/equipment")
@require_auth("equipment:read")
def list_equipment():
    return list_resource(equipment_config())


@bp.post("/equipment")
@require_auth("equipment:write")
def create_equipment():
    return create_resource(
        equipment_config(),
        location_prefix="/api/v1/equipment",
        required=("equipment_code",),
    )


@bp.get("/equipment/types")
@require_auth("equipment:read")
def equipment_types():
    from extensions import db
    from models.db_models import Equipment

    rows = db.session.query(Equipment.equipment_type).distinct().all()
    return ok([row[0] for row in rows if row[0]])


@bp.get("/equipment/<int:equipment_id>")
@require_auth("equipment:read")
def show_equipment(equipment_id):
    return show_resource(equipment_config(), equipment_id)


@bp.patch("/equipment/<int:equipment_id>")
@require_auth("equipment:write")
def patch_equipment(equipment_id):
    return update_resource(equipment_config(), equipment_id, partial=True)


@bp.put("/equipment/<int:equipment_id>")
@require_auth("equipment:write")
def put_equipment(equipment_id):
    return update_resource(equipment_config(), equipment_id, partial=False)


@bp.delete("/equipment/<int:equipment_id>")
@require_auth("equipment:write")
def remove_equipment(equipment_id):
    return delete_resource(equipment_config(), equipment_id)


# ---------------------------------------------------------------------------
# Regional manager -> state mappings
# ---------------------------------------------------------------------------


def manager_state_config():
    from models.db_models import RegionalManagerState

    return ResourceConfig(
        model=RegionalManagerState,
        table="regional_manager_states",
        spec_name="manager_state",
        searchable=("manager_email", "state"),
        sortable=("id", "manager_email", "state"),
        default_sort="manager_email",
    )


@bp.get("/regional-manager-states")
@require_auth("manager_states:read")
def list_manager_states():
    return list_resource(manager_state_config())


@bp.post("/regional-manager-states")
@require_auth("manager_states:write")
def create_manager_state():
    return create_resource(
        manager_state_config(),
        location_prefix="/api/v1/regional-manager-states",
        required=("manager_email", "state"),
    )


@bp.get("/regional-manager-states/by-manager/<path:email>")
@require_auth("manager_states:read")
def states_by_manager(email):
    """The scoping list every regional endpoint uses (routes/regional.py:36-49)."""
    from models.db_models import RegionalManagerState

    rows = (
        base_query(RegionalManagerState, wants_deleted())
        .filter(RegionalManagerState.manager_email.ilike(email))
        .order_by(RegionalManagerState.state)
        .all()
    )
    return ok({"manager_email": email, "states": [row.state for row in rows]})


@bp.put("/regional-manager-states/by-manager/<path:email>")
@require_auth("manager_states:write")
def replace_states_for_manager(email):
    """Replace a manager's whole state set in one transaction."""
    from extensions import db
    from models.db_models import RegionalManagerState

    body = json_body()
    states = body.get("states")
    if not isinstance(states, list):
        raise ValidationError(
            message="'states' must be a list.",
            details=[{"field": "states", "issue": "must be a list"}],
        )

    cleaned = sorted({str(s).strip() for s in states if str(s).strip()})

    RegionalManagerState.query.filter(
        RegionalManagerState.manager_email.ilike(email)
    ).delete(synchronize_session=False)

    for state in cleaned:
        db.session.add(RegionalManagerState(manager_email=email, state=state))
    commit("replace")

    return ok({"manager_email": email, "states": cleaned})


@bp.get("/regional-manager-states/<int:mapping_id>")
@require_auth("manager_states:read")
def show_manager_state(mapping_id):
    return show_resource(manager_state_config(), mapping_id)


@bp.patch("/regional-manager-states/<int:mapping_id>")
@require_auth("manager_states:write")
def patch_manager_state(mapping_id):
    return update_resource(manager_state_config(), mapping_id, partial=True)


@bp.delete("/regional-manager-states/<int:mapping_id>")
@require_auth("manager_states:write")
def remove_manager_state(mapping_id):
    return delete_resource(manager_state_config(), mapping_id)

# ==========================================================================
# res_alerts
# /api/v1/alerts - the missed-survey engine.
#
# ``GET /alerts`` NEVER writes.  This matters: the portal evaluates and commits as
# a side effect of rendering ``GET /admin``, so a machine client polling the
# equivalent read endpoint would be mutating production state on every poll.
# Writing is confined to ``POST /alerts/evaluate`` and ``POST /jobs/tick``.
# ==========================================================================

from flask import Blueprint, request

from core.engine import require_auth
from core.engine import json_body, list_resource
from core.engine import ok
from core.engine import ValidationError
from core.engine import compute_alerts, evaluate, missed_count, run_missed_engine


VARIANTS = ("admin", "regional")


def _states():
    raw = request.args.get("states")
    if not raw:
        return None
    return [s.strip() for s in raw.split(",") if s.strip()]


def _variant(value=None):
    variant = value or request.args.get("variant") or "regional"
    if variant not in VARIANTS:
        raise ValidationError(
            message=f"Unknown variant '{variant}'.",
            details=[{"field": "variant", "issue": "unknown", "value": variant,
                      "allowed": list(VARIANTS)}],
        )
    return variant


@bp.get("/alerts")
@require_auth("alerts:read")
def list_alerts():
    """
    Current missed-survey alerts. Read-only.

    ``?variant=admin`` reproduces the admin dashboard's behaviour exactly,
    including its loop-variable bug at routes/admin.py:521 where only one of
    several assignments gets marked.  ``regional`` (default) does what the logic
    clearly intends.

    ``changes`` shows what a write run *would* do.
    """
    result = evaluate(states=_states(), variant=_variant())
    result["committed"] = False
    result["missed_total"] = missed_count(_states())
    return ok(result)


@bp.post("/alerts/evaluate")
@require_auth("alerts:write")
def evaluate_alerts():
    """Run the engine for real and persist the status changes."""
    body = json_body(required=False)
    result = run_missed_engine(
        states=body.get("states") or _states(),
        variant=_variant(body.get("variant")),
        dry_run=bool(body.get("dry_run")),
    )
    result["missed_total"] = missed_count(body.get("states") or _states())
    return ok(result)


@bp.get("/alerts/missed")
@require_auth("alerts:read")
def missed_assignments():
    """
    The "Missed Surveys" table - routes/admin.py:573-578.

    Ordered by day then section, as the page does.
    """
    from models.db_models import SurveyAssignment

    from core.engine import base_query
    from core.endpoints import assignment_config

    query = base_query(SurveyAssignment).filter(SurveyAssignment.status == "missed")
    return list_resource(assignment_config(), query=query)


@bp.get("/alerts/summary")
@require_auth("alerts:read")
def alerts_summary():
    """Counts by state and by day, for a KPI strip."""
    from extensions import db
    from models.db_models import SurveyAssignment

    states = _states()
    query = SurveyAssignment.query.filter_by(status="missed")
    if states:
        query = query.filter(SurveyAssignment.state.in_(states))

    by_state = dict(
        query.with_entities(SurveyAssignment.state, db.func.count(SurveyAssignment.id))
        .group_by(SurveyAssignment.state)
        .all()
    )
    by_day = dict(
        query.with_entities(
            SurveyAssignment.survey_day, db.func.count(SurveyAssignment.id)
        )
        .group_by(SurveyAssignment.survey_day)
        .all()
    )

    live = compute_alerts(states=states)

    return ok(
        {
            "missed_total": query.count(),
            "by_state": {k: int(v) for k, v in by_state.items() if k},
            "by_day": {k: int(v) for k, v in by_day.items() if k},
            "unacknowledged_alerts": len(live),
            "acknowledged": SurveyAssignment.query.filter_by(
                status="missed", alert_acknowledged=True
            ).count(),
        }
    )

# ==========================================================================
# res_media
# /api/v1/media - uploads, and the media fields on a survey.
#
# Uses the portal's own Drive/Supabase helpers, so files land in the same folders
# with the same naming and the same public-read permission.
# ==========================================================================

from flask import Blueprint, request

from core.engine import require_auth
from core.config import MEDIA_FIELDS
from core.engine import commit, get_or_404, json_body, wants_deleted
from core.engine import no_content, ok
from core.engine import ValidationError
from core.engine import Options, spec
from core.engine import (
    media_backends_status,
    upload_image,
    upload_pdf,
    upload_pdf_to_supabase,
)



@bp.get("/media/status")
@require_auth("media:write")
def media_status():
    return ok(media_backends_status())


@bp.post("/media/images")
@require_auth("media:write")
def post_image():
    """
    Upload a dashcam photo. multipart ``file``.

    ``?compress=false`` skips the Pillow resize the portal normally applies.
    """
    compress = request.args.get("compress", "true").lower() not in ("0", "false", "no")
    result = upload_image(request.files.get("file"), compress=compress)
    return ok(
        {
            "id": result.get("id"),
            "view_url": result.get("view_url"),
            "image_url": result.get("image_url"),
            "backend": "google_drive",
        },
        status=201,
    )


@bp.post("/media/pdfs")
@require_auth("media:write")
def post_pdf():
    """Upload a PDF to Drive. multipart ``file``; ``?optimize=false`` to skip pikepdf."""
    optimize = request.args.get("optimize", "true").lower() not in ("0", "false", "no")
    result = upload_pdf(
        request.files.get("file"),
        optimize=optimize,
        filename=request.args.get("filename"),
    )
    return ok(
        {
            "id": result.get("id"),
            "view_url": result.get("view_url"),
            "backend": "google_drive",
        },
        status=201,
    )


@bp.post("/media/pdfs/supabase")
@require_auth("media:write")
def post_pdf_supabase():
    """Upload a PDF to the Supabase bucket instead of Drive."""
    return ok(upload_pdf_to_supabase(request.files.get("file")), status=201)


@bp.put("/surveys/<int:survey_id>/media/<kind>")
@require_auth("surveys:write")
def set_survey_media(survey_id, kind):
    """
    Point a survey's media field at a URL, or upload the file inline.

    ``kind`` is one of dashcam_photo, end_survey_pdf, end_survey_photo,
    settings_photo.
    """
    from models.db_models import Survey

    if kind not in MEDIA_FIELDS:
        raise ValidationError(
            message=f"Unknown media field '{kind}'.",
            details=[{"field": "kind", "issue": "unknown", "value": kind,
                      "allowed": list(MEDIA_FIELDS)}],
        )

    survey = get_or_404(Survey, survey_id, wants_deleted())

    if request.files.get("file"):
        storage = request.files["file"]
        if kind == "end_survey_pdf":
            from core.engine import upload_survey_pdf

            url = upload_survey_pdf(storage, survey)["view_url"]
        else:
            url = upload_image(storage)["image_url"]
    else:
        url = json_body().get("url")
        if not url:
            raise ValidationError(
                message="Provide 'url' or upload a multipart 'file'.",
                details=[{"field": "url", "issue": "required"}],
            )

    setattr(survey, kind, url)
    commit("media")
    return ok(spec("survey").serialize(survey, Options.from_request()))


@bp.delete("/surveys/<int:survey_id>/media/<kind>")
@require_auth("surveys:write")
def clear_survey_media(survey_id, kind):
    """Clears the URL on the record. The file itself is left in Drive."""
    from models.db_models import Survey

    if kind not in MEDIA_FIELDS:
        raise ValidationError(
            message=f"Unknown media field '{kind}'.",
            details=[{"field": "kind", "issue": "unknown", "value": kind}],
        )

    survey = get_or_404(Survey, survey_id, wants_deleted())
    setattr(survey, kind, None)
    commit("media")
    return no_content()

# ==========================================================================
# res_reports
# /api/v1/reports and /api/v1/exports.
#
# The reports page aggregates in Python (routes/admin.py:872-898) and the Excel
# export rebuilds the same query with the same filters (routes/admin.py:937-1120).
# Both are reproduced here, plus a few rollups the portal does not have but the
# data supports.
# ==========================================================================

from io import BytesIO

from flask import Blueprint, Response, request, send_file

from core.engine import require_auth
from core.config import EXPORT_COLUMNS, SURVEY_COMPLETED, TEAMS, team_display
from core.engine import base_query, wants_deleted
from core.config import report_summary, survey_duration_minutes, team_km_totals
from core.engine import ok
from core.engine import ValidationError
from core.engine import ListParams, parse_int
from core.config import utc_now
from core.config import week_list, week_window



def _filtered_surveys(default_completed=True):
    """Apply the shared survey filter set; defaults to completed, as the page does."""
    from models.db_models import Survey

    from core.endpoints import survey_config

    cfg = survey_config()
    params = ListParams.from_request(cfg, extra_allowed={"format"})
    query = base_query(Survey, wants_deleted())

    if default_completed and not request.args.get("status") and not request.args.get(
        "status__in"
    ):
        query = query.filter(Survey.status == SURVEY_COMPLETED)

    return params.apply_filters(query), params


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------


@bp.get("/reports/summary")
@require_auth("reports:read")
def summary():
    """
    The four stat cards on /reports.

    ``completed_surveys``, ``total_km``, ``captains`` (distinct), ``total_hours``
    - exactly routes/admin.py:872-898.
    """
    query, params = _filtered_surveys()
    surveys = query.all()
    data = report_summary(surveys)
    data["filters_applied"] = params.filters_applied
    return ok(data, meta=params.meta())


@bp.get("/reports/team-km")
@require_auth("reports:read")
def team_km():
    """
    KM per team - the second KPI row on the admin dashboard.

    Only completed surveys count, and a NULL section_length counts as zero, as
    routes/admin.py:258-290 does.
    """
    query, params = _filtered_surveys(default_completed=False)
    surveys = query.all()
    totals = team_km_totals(surveys)

    return ok(
        {
            "teams": [
                {
                    "key": key,
                    "display": team_display(key),
                    "states": list(meta["states"]),
                    "km": totals[key],
                    "surveys": sum(
                        1
                        for s in surveys
                        if s.status == SURVEY_COMPLETED and s.state in meta["states"]
                    ),
                }
                for key, meta in TEAMS.items()
            ],
            "total_km": totals["total"],
        },
        meta=params.meta(),
    )


@bp.get("/reports/weekly")
@require_auth("reports:read")
def weekly():
    """Per-project-week rollup."""
    from models.db_models import Survey

    from_week = parse_int(request.args.get("from_week", 1), "from_week")
    to_week = request.args.get("to_week")
    weeks = week_list()
    to_week = parse_int(to_week, "to_week") if to_week else (weeks[-1] if weeks else 1)

    if from_week > to_week:
        raise ValidationError(
            message="'from_week' must be <= 'to_week'.",
            details=[{"field": "from_week", "issue": "greater than to_week"}],
        )

    rows = []
    for week in range(from_week, to_week + 1):
        start, end = week_window(week)
        surveys = (
            base_query(Survey)
            .filter(
                Survey.status == SURVEY_COMPLETED,
                Survey.start_time >= start,
                Survey.start_time < end,
            )
            .all()
        )
        data = report_summary(surveys)
        data.update(
            {
                "week": week,
                "start_raw": start.isoformat(),
                "end_raw": end.isoformat(),
                "team_km": team_km_totals(surveys),
            }
        )
        rows.append(data)

    return ok({"weeks": rows})


@bp.get("/reports/captain-performance")
@require_auth("reports:read")
def captain_performance():
    """Per-captain totals, including their current missed-assignment count."""
    from models.db_models import SurveyAssignment

    query, params = _filtered_surveys()
    surveys = query.all()

    by_captain = {}
    for survey in surveys:
        bucket = by_captain.setdefault(
            survey.captain_email,
            {
                "captain_email": survey.captain_email,
                "captain_name": survey.captain_name,
                "surveys": 0,
                "km": 0.0,
                "minutes": 0.0,
                "upload_minutes": [],
            },
        )
        bucket["surveys"] += 1
        bucket["km"] += survey.section_length or 0
        duration = survey_duration_minutes(survey)
        if duration:
            bucket["minutes"] += duration
        if survey.video_pending_start_time and survey.video_upload_time:
            bucket["upload_minutes"].append(
                (survey.video_upload_time - survey.video_pending_start_time).total_seconds()
                / 60
            )

    missed = dict(
        SurveyAssignment.query.with_entities(
            SurveyAssignment.captain_email, SurveyAssignment.status
        )
        .filter(SurveyAssignment.status == "missed")
        .all()
    )

    rows = []
    for email, bucket in by_captain.items():
        uploads = bucket.pop("upload_minutes")
        rows.append(
            {
                **bucket,
                "km": round(bucket["km"], 2),
                "hours": round(bucket["minutes"] / 60, 2),
                "avg_upload_minutes": round(sum(uploads) / len(uploads), 1)
                if uploads
                else None,
                "missed_assignments": 1 if email in missed else 0,
            }
        )
        rows[-1].pop("minutes", None)

    rows.sort(key=lambda r: -r["surveys"])
    return ok({"captains": rows}, meta=params.meta())


@bp.get("/reports/upload-sla")
@require_auth("reports:read")
def upload_sla():
    """
    Video-upload turnaround against the dashboards' 480-minute red line.

    Percentiles are computed over completed uploads only.
    """
    from core.config import UPLOAD_DURATION_RED_MINUTES

    threshold = parse_int(
        request.args.get("threshold", UPLOAD_DURATION_RED_MINUTES), "threshold"
    )
    query, params = _filtered_surveys(default_completed=False)

    rows = []
    for survey in query.all():
        if not (survey.video_pending_start_time and survey.video_upload_time):
            continue
        minutes = int(
            (survey.video_upload_time - survey.video_pending_start_time).total_seconds() / 60
        )
        rows.append(
            {
                "survey_id": survey.id,
                "section_no": survey.section_no,
                "captain_name": survey.captain_name,
                "state": survey.state,
                "minutes": minutes,
                "over_threshold": minutes > threshold,
            }
        )

    values = sorted(r["minutes"] for r in rows)

    def percentile(fraction):
        if not values:
            return None
        index = min(int(len(values) * fraction), len(values) - 1)
        return values[index]

    return ok(
        {
            "threshold_minutes": threshold,
            "count": len(rows),
            "breach_count": sum(1 for r in rows if r["over_threshold"]),
            "p50": percentile(0.50),
            "p90": percentile(0.90),
            "p99": percentile(0.99),
            "max": values[-1] if values else None,
            "surveys": sorted(rows, key=lambda r: -r["minutes"])[:200],
        },
        meta=params.meta(),
    )


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------


def _export_rows(surveys):
    for survey in surveys:
        yield [
            survey.start_time.strftime("%d-%m-%Y") if survey.start_time else "",
            survey.cycle_no,
            survey.captain_name,
            survey.state,
            survey.section_no,
            survey.upc_code,
            survey.stretch_code,
            survey.section_length,
            survey.survey_type,
            survey.status,
        ]


@bp.get("/exports/surveys.xlsx")
@require_auth("exports:read")
def export_surveys_xlsx():
    """
    Byte-compatible with the portal's ``Survey_Report.xlsx``.

    Same columns, same order, same ``start_time DESC`` ordering
    (routes/admin.py:937-1120).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    from models.db_models import Survey

    query, _params = _filtered_surveys()
    surveys = query.order_by(Survey.start_time.desc()).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Survey Report"

    for column_index, header in enumerate(EXPORT_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.value = header
        cell.font = Font(bold=True)

    for row_index, values in enumerate(_export_rows(surveys), start=2):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index).value = value

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Survey_Report.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


@bp.get("/exports/surveys.csv")
@require_auth("exports:read")
def export_surveys_csv():
    import csv
    import io

    from models.db_models import Survey

    query, _params = _filtered_surveys()
    surveys = query.order_by(Survey.start_time.desc()).all()

    def generate():
        buffer = io.StringIO()
        writer = csv.writer(buffer, lineterminator="\n")
        writer.writerow(EXPORT_COLUMNS)
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)
        for values in _export_rows(surveys):
            writer.writerow(values)
            yield buffer.getvalue()
            buffer.seek(0)
            buffer.truncate(0)

    stamp = utc_now().strftime("%Y%m%d")
    return Response(
        generate(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="Survey_Report_{stamp}.csv"'
        },
    )


@bp.get("/exports/assignments.xlsx")
@require_auth("exports:read")
def export_assignments_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font

    from models.db_models import SurveyAssignment

    from core.endpoints import assignment_config

    cfg = assignment_config()
    params = ListParams.from_request(cfg)
    query = params.apply(base_query(SurveyAssignment, wants_deleted()))
    rows = query.all()

    headers = [
        "Section No", "UPC Code", "Stretch", "State", "Captain",
        "Survey Day", "Survey Type", "KM", "Status", "Enabled",
        "Dashcam", "Powerbank", "NH", "RO", "PIU",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Assignments"
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.value = header
        cell.font = Font(bold=True)

    for row_index, item in enumerate(rows, start=2):
        for column_index, value in enumerate(
            [
                item.section_no, item.upc_code, item.stretch_code, item.state,
                item.main_person, item.survey_day, item.survey_type,
                item.section_length, item.status, bool(item.survey_enabled),
                item.dashcam_code, item.powerbank_code, item.nh_number,
                item.ro, item.piu,
            ],
            start=1,
        ):
            sheet.cell(row=row_index, column=column_index).value = value

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="Survey_Assignments.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


@bp.get("/exports/columns")
@require_auth("exports:read")
def export_columns():
    return ok({"surveys": list(EXPORT_COLUMNS)})

# ==========================================================================
# res_audit
# /api/v1/audit - who logged in, from where, and everything they did.
#
# Three feeds:
#
#   /audit/logins    every login success, login failure and logout, with IP,
#                    device, geo and the session id that ties it to later requests
#   /audit/requests  the full access log - every request, portal and API
#   /audit/sessions  logins collapsed into sessions, with their request counts
#
# Plus rollups (``/audit/summary``, ``/audit/active-users``, ``/audit/ips``) so the
# other site can render dashboards without pulling raw rows.
#
# Everything here needs the ``audit:read`` scope, which is NOT part of the
# read-only bundle - access logs are more sensitive than survey data, so a mirror
# token has to be granted it deliberately.
# ==========================================================================

from datetime import datetime, timedelta

from flask import Blueprint, request

from core.engine import require_auth
from core.engine import ok, paginated
from core.engine import NotFound, ValidationError
from core.engine import ListParams, ResourceConfig, parse_int
from core.engine import Options, jsonify_value
from core.config import UTC_WALL, as_ist, as_utc, iso, utc_now


LOGIN_EVENT_TYPES = ("login_success", "login_failed", "logout", "api_token_used")


def _login_config():
    from core.models import LoginEvent

    return ResourceConfig(
        model=LoginEvent,
        table="api_login_events",
        spec_name="login_event",
        searchable=("email", "username", "name", "ip", "login_identifier", "user_agent"),
        sortable=(
            "id",
            "created_at",
            "event_type",
            "email",
            "role",
            "ip",
            "country",
            "user_id",
        ),
        default_sort="-created_at",
    )


def _request_config():
    from core.models import RequestLog

    return ResourceConfig(
        model=RequestLog,
        table="api_request_logs",
        spec_name="request_log",
        searchable=("path", "email", "ip", "user_agent", "token_sub"),
        sortable=(
            "id",
            "created_at",
            "path",
            "method",
            "status_code",
            "duration_ms",
            "channel",
            "user_id",
            "ip",
        ),
        default_sort="-created_at",
    )


def _audit_row(row, table):
    """
    Audit rows are plain records - serialize by column, expanding timestamps.

    Not routed through core.serializers.spec() because these models are not part
    of the portal domain and need no FieldSpec.
    """
    data = {}
    for column in row.__table__.columns:
        value = getattr(row, column.key)
        if column.key == "created_at":
            data["created_at_utc"] = iso(as_utc(value, UTC_WALL))
            data["created_at_ist"] = iso(as_ist(value, UTC_WALL))
            data["created_at_raw"] = value.isoformat() if value else None
        else:
            data[column.key] = jsonify_value(value)
    _ = table
    return data


def _window_filter(query, model):
    """Shared ``?since=`` / ``?until=`` / ``?days=`` window on created_at."""
    since = request.args.get("since")
    until = request.args.get("until")
    days = request.args.get("days")

    if days:
        cutoff = utc_now() - timedelta(days=parse_int(days, "days"))
        query = query.filter(model.created_at >= cutoff)
    if since:
        query = query.filter(model.created_at >= _parse_instant(since, "since"))
    if until:
        query = query.filter(model.created_at < _parse_instant(until, "until"))
    return query


def _parse_instant(value, field):
    from core.config import TimeParseError, parse_datetime

    try:
        return parse_datetime(value, UTC_WALL, assume_tz="UTC")
    except TimeParseError as exc:
        raise ValidationError(
            message=str(exc),
            details=[{"field": field, "issue": "invalid datetime", "value": value}],
        )


# ---------------------------------------------------------------------------
# Login events
# ---------------------------------------------------------------------------


@bp.get("/audit/logins")
@require_auth("audit:read")
def list_logins():
    """
    Every authentication event.

    Filters: ``event_type``, ``email``, ``user_id``, ``role``, ``ip``, ``country``,
    ``device_type``, plus ``since`` / ``until`` / ``days`` and ``q``.
    """
    from core.models import LoginEvent

    config = _login_config()
    params = ListParams.from_request(config, extra_allowed={"since", "until", "days"})
    query = _window_filter(LoginEvent.query, LoginEvent)
    query = params.apply(query)
    rows, page_info = params.paginate(query)
    return paginated(
        [_audit_row(row, "api_login_events") for row in rows],
        page_info,
        meta=params.meta(),
    )


@bp.get("/audit/logins/<int:event_id>")
@require_auth("audit:read")
def show_login(event_id):
    from core.models import LoginEvent

    row = LoginEvent.query.get(event_id)
    if not row:
        raise NotFound(message=f"Login event {event_id} not found.")
    return ok(_audit_row(row, "api_login_events"))


def login_events_for_user(user):
    """Backing implementation for ``GET /users/{id}/logins``."""
    from core.models import LoginEvent

    config = _login_config()
    params = ListParams.from_request(config, extra_allowed={"since", "until", "days"})
    query = LoginEvent.query.filter(
        (LoginEvent.user_id == user.id) | (LoginEvent.email == user.email)
    )
    query = _window_filter(query, LoginEvent)
    query = params.apply(query)
    rows, page_info = params.paginate(query)
    return paginated(
        [_audit_row(row, "api_login_events") for row in rows],
        page_info,
        meta=params.meta(),
    )


@bp.get("/audit/failed-logins")
@require_auth("audit:read")
def failed_logins():
    """
    Failed attempts grouped by identifier and IP.

    The shape you want for spotting credential stuffing: repeated failures
    against one account, or one IP hitting many accounts.
    """
    from extensions import db

    from core.models import LoginEvent

    days = parse_int(request.args.get("days", 7), "days")
    cutoff = utc_now() - timedelta(days=days)

    base = LoginEvent.query.filter(
        LoginEvent.event_type == "login_failed", LoginEvent.created_at >= cutoff
    )

    by_identifier = (
        base.with_entities(
            LoginEvent.login_identifier,
            db.func.count(LoginEvent.id),
            db.func.count(db.distinct(LoginEvent.ip)),
            db.func.max(LoginEvent.created_at),
        )
        .group_by(LoginEvent.login_identifier)
        .order_by(db.func.count(LoginEvent.id).desc())
        .limit(100)
        .all()
    )

    by_ip = (
        base.with_entities(
            LoginEvent.ip,
            db.func.count(LoginEvent.id),
            db.func.count(db.distinct(LoginEvent.login_identifier)),
            db.func.max(LoginEvent.created_at),
        )
        .group_by(LoginEvent.ip)
        .order_by(db.func.count(LoginEvent.id).desc())
        .limit(100)
        .all()
    )

    return ok(
        {
            "window_days": days,
            "total_failures": base.count(),
            "by_identifier": [
                {
                    "login_identifier": identifier,
                    "failures": int(count),
                    "distinct_ips": int(ips),
                    "last_attempt_utc": iso(as_utc(last, UTC_WALL)),
                }
                for identifier, count, ips, last in by_identifier
            ],
            "by_ip": [
                {
                    "ip": ip,
                    "failures": int(count),
                    "distinct_identifiers": int(identifiers),
                    "last_attempt_utc": iso(as_utc(last, UTC_WALL)),
                }
                for ip, count, identifiers, last in by_ip
            ],
        }
    )


# ---------------------------------------------------------------------------
# Request log
# ---------------------------------------------------------------------------


@bp.get("/audit/requests")
@require_auth("audit:read")
def list_requests():
    """
    The full access log.

    Filters: ``path``, ``path__like``, ``method``, ``status_code``,
    ``status_code__gte``, ``channel``, ``user_id``, ``email``, ``ip``,
    ``session_id``, ``token_sub``, plus the time window and ``q``.
    """
    from core.models import RequestLog

    config = _request_config()
    params = ListParams.from_request(config, extra_allowed={"since", "until", "days"})
    query = _window_filter(RequestLog.query, RequestLog)
    query = params.apply(query)
    rows, page_info = params.paginate(query)
    return paginated(
        [_audit_row(row, "api_request_logs") for row in rows],
        page_info,
        meta=params.meta(),
    )


@bp.get("/audit/requests/<int:log_id>")
@require_auth("audit:read")
def show_request(log_id):
    from core.models import RequestLog

    row = RequestLog.query.get(log_id)
    if not row:
        raise NotFound(message=f"Request log {log_id} not found.")
    return ok(_audit_row(row, "api_request_logs"))


# ---------------------------------------------------------------------------
# Sessions
# ---------------------------------------------------------------------------


@bp.get("/audit/sessions")
@require_auth("audit:read")
def list_sessions():
    """
    Logins rolled up into sessions.

    Each row is one successful login plus what happened on that session id:
    request count, last activity, and the logout time if there was one.
    """
    from extensions import db

    from core.models import LoginEvent, RequestLog

    days = parse_int(request.args.get("days", 30), "days")
    limit = min(parse_int(request.args.get("limit", 100), "limit"), 500)
    cutoff = utc_now() - timedelta(days=days)

    logins = (
        LoginEvent.query.filter(
            LoginEvent.event_type == "login_success", LoginEvent.created_at >= cutoff
        )
        .order_by(LoginEvent.created_at.desc())
        .limit(limit)
        .all()
    )

    session_ids = [row.session_id for row in logins if row.session_id]
    activity = {}
    if session_ids:
        rows = (
            RequestLog.query.with_entities(
                RequestLog.session_id,
                db.func.count(RequestLog.id),
                db.func.max(RequestLog.created_at),
            )
            .filter(RequestLog.session_id.in_(session_ids))
            .group_by(RequestLog.session_id)
            .all()
        )
        activity = {sid: (int(count), last) for sid, count, last in rows}

    logouts = {}
    if session_ids:
        for row in LoginEvent.query.filter(
            LoginEvent.event_type == "logout", LoginEvent.session_id.in_(session_ids)
        ).all():
            logouts[row.session_id] = row.created_at

    sessions = []
    for login in logins:
        count, last_seen = activity.get(login.session_id, (0, None))
        logged_out = logouts.get(login.session_id)
        sessions.append(
            {
                "session_id": login.session_id,
                "user_id": login.user_id,
                "name": login.name,
                "email": login.email,
                "role": login.role,
                "ip": login.ip,
                "device_type": login.device_type,
                "browser": login.browser,
                "platform": login.platform,
                "country": login.country,
                "city": login.city,
                "login_at_utc": iso(as_utc(login.created_at, UTC_WALL)),
                "login_at_ist": iso(as_ist(login.created_at, UTC_WALL)),
                "last_seen_utc": iso(as_utc(last_seen, UTC_WALL)) if last_seen else None,
                "logout_at_utc": iso(as_utc(logged_out, UTC_WALL)) if logged_out else None,
                "request_count": count,
                "duration_minutes": (
                    int(((logged_out or last_seen) - login.created_at).total_seconds() / 60)
                    if (logged_out or last_seen)
                    else None
                ),
                "ended": bool(logged_out),
            }
        )

    return ok({"window_days": days, "sessions": sessions})


@bp.get("/audit/sessions/<session_id>")
@require_auth("audit:read")
def show_session(session_id):
    """Everything that happened on one session, in order."""
    from core.models import LoginEvent, RequestLog

    events = (
        LoginEvent.query.filter_by(session_id=session_id)
        .order_by(LoginEvent.created_at)
        .all()
    )
    requests_made = (
        RequestLog.query.filter_by(session_id=session_id)
        .order_by(RequestLog.created_at)
        .limit(1000)
        .all()
    )
    if not events and not requests_made:
        raise NotFound(message=f"No activity for session '{session_id}'.")

    return ok(
        {
            "session_id": session_id,
            "auth_events": [_audit_row(row, "api_login_events") for row in events],
            "requests": [_audit_row(row, "api_request_logs") for row in requests_made],
            "request_count": len(requests_made),
        }
    )


# ---------------------------------------------------------------------------
# Rollups
# ---------------------------------------------------------------------------


@bp.get("/audit/summary")
@require_auth("audit:read")
def audit_summary():
    from extensions import db

    from core.models import LoginEvent, RequestLog

    days = parse_int(request.args.get("days", 7), "days")
    cutoff = utc_now() - timedelta(days=days)

    logins = LoginEvent.query.filter(LoginEvent.created_at >= cutoff)
    requests_made = RequestLog.query.filter(RequestLog.created_at >= cutoff)

    by_type = dict(
        logins.with_entities(LoginEvent.event_type, db.func.count(LoginEvent.id))
        .group_by(LoginEvent.event_type)
        .all()
    )
    by_role = dict(
        logins.filter(LoginEvent.event_type == "login_success")
        .with_entities(LoginEvent.role, db.func.count(LoginEvent.id))
        .group_by(LoginEvent.role)
        .all()
    )
    by_device = dict(
        logins.filter(LoginEvent.event_type == "login_success")
        .with_entities(LoginEvent.device_type, db.func.count(LoginEvent.id))
        .group_by(LoginEvent.device_type)
        .all()
    )
    by_channel = dict(
        requests_made.with_entities(RequestLog.channel, db.func.count(RequestLog.id))
        .group_by(RequestLog.channel)
        .all()
    )
    by_status = dict(
        requests_made.with_entities(RequestLog.status_code, db.func.count(RequestLog.id))
        .group_by(RequestLog.status_code)
        .all()
    )

    slowest = (
        requests_made.filter(RequestLog.duration_ms.isnot(None))
        .order_by(RequestLog.duration_ms.desc())
        .limit(10)
        .all()
    )

    return ok(
        {
            "window_days": days,
            "logins": {
                "by_event_type": {k: int(v) for k, v in by_type.items() if k},
                "by_role": {k: int(v) for k, v in by_role.items() if k},
                "by_device": {k: int(v) for k, v in by_device.items() if k},
                "distinct_users": logins.filter(LoginEvent.user_id.isnot(None))
                .with_entities(db.func.count(db.distinct(LoginEvent.user_id)))
                .scalar()
                or 0,
                "distinct_ips": logins.with_entities(
                    db.func.count(db.distinct(LoginEvent.ip))
                ).scalar()
                or 0,
            },
            "requests": {
                "total": requests_made.count(),
                "by_channel": {k: int(v) for k, v in by_channel.items() if k},
                "by_status": {str(k): int(v) for k, v in by_status.items() if k},
                "errors": requests_made.filter(RequestLog.status_code >= 400).count(),
                "slowest": [
                    {
                        "path": row.path,
                        "method": row.method,
                        "duration_ms": row.duration_ms,
                        "status_code": row.status_code,
                        "at_utc": iso(as_utc(row.created_at, UTC_WALL)),
                    }
                    for row in slowest
                ],
            },
        }
    )


@bp.get("/audit/active-users")
@require_auth("audit:read")
def active_users():
    """Who has been using the portal, most recent first."""
    from extensions import db

    from core.models import LoginEvent

    days = parse_int(request.args.get("days", 30), "days")
    cutoff = utc_now() - timedelta(days=days)

    rows = (
        LoginEvent.query.filter(
            LoginEvent.event_type == "login_success", LoginEvent.created_at >= cutoff
        )
        .with_entities(
            LoginEvent.user_id,
            LoginEvent.email,
            LoginEvent.name,
            LoginEvent.role,
            db.func.count(LoginEvent.id),
            db.func.max(LoginEvent.created_at),
            db.func.count(db.distinct(LoginEvent.ip)),
        )
        .group_by(LoginEvent.user_id, LoginEvent.email, LoginEvent.name, LoginEvent.role)
        .order_by(db.func.max(LoginEvent.created_at).desc())
        .all()
    )

    return ok(
        {
            "window_days": days,
            "users": [
                {
                    "user_id": user_id,
                    "email": email,
                    "name": name,
                    "role": role,
                    "login_count": int(count),
                    "distinct_ips": int(ips),
                    "last_login_utc": iso(as_utc(last, UTC_WALL)),
                    "last_login_ist": iso(as_ist(last, UTC_WALL)),
                }
                for user_id, email, name, role, count, last, ips in rows
            ],
        }
    )


@bp.get("/audit/ips")
@require_auth("audit:read")
def ip_activity():
    """Activity grouped by source IP, with the accounts seen from each."""
    from extensions import db

    from core.models import LoginEvent

    days = parse_int(request.args.get("days", 30), "days")
    cutoff = utc_now() - timedelta(days=days)

    # Failed logins have no matched user, so count the typed identifier too -
    # otherwise an IP that only ever fails looks like it touched zero accounts,
    # which is exactly backwards for spotting an attack.
    account = db.func.coalesce(LoginEvent.email, LoginEvent.login_identifier)

    rows = (
        LoginEvent.query.filter(LoginEvent.created_at >= cutoff)
        .with_entities(
            LoginEvent.ip,
            LoginEvent.country,
            LoginEvent.city,
            db.func.count(LoginEvent.id),
            db.func.count(db.distinct(account)),
            db.func.max(LoginEvent.created_at),
        )
        .group_by(LoginEvent.ip, LoginEvent.country, LoginEvent.city)
        .order_by(db.func.count(LoginEvent.id).desc())
        .limit(200)
        .all()
    )

    return ok(
        {
            "window_days": days,
            "ips": [
                {
                    "ip": ip,
                    "country": country,
                    "city": city,
                    "event_count": int(count),
                    "distinct_accounts": int(accounts),
                    "last_seen_utc": iso(as_utc(last, UTC_WALL)),
                }
                for ip, country, city, count, accounts, last in rows
            ],
        }
    )


@bp.get("/audit/config")
@require_auth("audit:read")
def audit_config():
    """What is being captured and for how long."""
    from core.engine import (
        anonymise_ip,
        audit_enabled,
        request_log_mode,
        retention_days,
    )

    return ok(
        {
            "enabled": audit_enabled(),
            "request_log_mode": request_log_mode(),
            "anonymise_ip": anonymise_ip(),
            "retention_days": retention_days(),
            "login_event_types": list(LOGIN_EVENT_TYPES),
            "note": (
                "IP addresses are personal data. Retention is bounded by "
                "AUDIT_RETENTION_DAYS and purged by POST /api/v1/jobs/tick."
            ),
        }
    )


_ = (Options, datetime)

# ==========================================================================
# res_dump
# /api/v1/dump - pull the entire database.
#
# Built for a nightly cron on another machine.  The whole DB is fetchable in one
# request, per table, or incrementally.
#
# Design notes
# ------------
# *Streaming, not buffering.*  ``surveys`` has ~50 columns and grows without bound;
# building a full JSON document in memory would eventually exhaust a serverless
# function.  Every dump endpoint streams NDJSON (one JSON object per line) using a
# server-side cursor, so memory stays flat regardless of table size.
#
# *Resumable.*  ``?after_id=`` continues from a known row, so a dump interrupted by
# a timeout can be resumed rather than restarted.
#
# *Verifiable.*  ``/dump/manifest`` gives row counts and a content checksum per
# table, so the receiving end can prove the dump landed intact.
#
# *Incremental.*  ``?since=`` filters to rows whose id is newer, and
# ``/dump/changes`` replays the change log, so a daily job can pull only deltas
# once the first full dump is in place.
#
# Formats: ``ndjson`` (default, streaming), ``json`` (buffered - only for small
# tables), ``csv``, ``sql`` (INSERT statements).  Add ``?compress=gzip`` for any of
# them.
# ==========================================================================

import csv
import gzip
import hashlib
import io
import json

from flask import Blueprint, Response, request, stream_with_context

from core.engine import require_auth
from core.engine import ok
from core.engine import Forbidden, ValidationError
from core.engine import parse_bool, parse_int
from core.engine import jsonify_value
from core.config import UTC_WALL, as_utc, iso, utc_now


#: Portal tables - the actual business data.
PORTAL_TABLES = {
    "users": ("models.db_models", "User"),
    "surveys": ("models.db_models", "Survey"),
    "survey_assignments": ("models.db_models", "SurveyAssignment"),
    "survey_schedule": ("models.db_models", "SurveySchedule"),
    "equipment": ("models.db_models", "Equipment"),
    "regional_manager_states": ("models.db_models", "RegionalManagerState"),
}

#: API-owned tables. Excluded by default - api_login_events and api_request_logs
#: contain IP addresses, and api_tokens contains credential metadata.
API_TABLES = {
    "api_login_events": ("core.api_models", "LoginEvent"),
    "api_request_logs": ("core.api_models", "RequestLog"),
    "change_log": ("core.api_models", "ChangeLog"),
    "api_deleted_records": ("core.api_models", "DeletedRecord"),
    "webhook_outbox": ("core.api_models", "WebhookOutbox"),
    "webhook_deliveries": ("core.api_models", "WebhookDelivery"),
    "webhook_endpoints": ("core.api_models", "WebhookEndpoint"),
    "api_tokens": ("core.api_models", "ApiToken"),
}

ALL_TABLES = {**PORTAL_TABLES, **API_TABLES}

#: Columns never written to a dump, whatever is requested.
REDACTED_COLUMNS = {
    "users": {"password_hash"},
    "webhook_endpoints": {"secret"},
    "api_tokens": {"jti"},
}

FORMATS = ("ndjson", "json", "csv", "sql")

#: Rows fetched per round-trip while streaming.
CHUNK_SIZE = 1000


def _resolve(table):
    if table not in ALL_TABLES:
        raise ValidationError(
            message=f"Unknown table '{table}'.",
            details=[
                {
                    "field": "table",
                    "issue": "unknown",
                    "value": table,
                    "allowed": sorted(ALL_TABLES),
                }
            ],
        )
    module_path, class_name = ALL_TABLES[table]
    import importlib

    return getattr(importlib.import_module(module_path), class_name)


def _requested_tables():
    """Which tables this request covers, honouring include/exclude and scopes."""
    from core.engine import current_client

    raw = request.args.get("tables")
    include_api = parse_bool(request.args.get("include_api_tables", "false"), "include_api_tables")
    include_audit = parse_bool(
        request.args.get("include_audit", "false"), "include_audit"
    )

    if raw:
        names = [name.strip() for name in raw.split(",") if name.strip()]
        unknown = [name for name in names if name not in ALL_TABLES]
        if unknown:
            raise ValidationError(
                message="Unknown table(s).",
                details=[
                    {"field": "tables", "issue": "unknown", "value": unknown,
                     "allowed": sorted(ALL_TABLES)}
                ],
            )
    else:
        names = list(PORTAL_TABLES)
        if include_api:
            names += [t for t in API_TABLES if t not in ("api_login_events", "api_request_logs")]
        if include_audit:
            names += ["api_login_events", "api_request_logs"]

    # Audit tables hold IP addresses; require the audit scope explicitly.
    client = current_client()
    if client and any(name in ("api_login_events", "api_request_logs") for name in names):
        client.require("audit:read")
    if client and "api_tokens" in names:
        client.require("admin:destroy")

    return names


def _columns(model, table):
    redacted = REDACTED_COLUMNS.get(table, set())
    return [c.key for c in model.__table__.columns if c.key not in redacted]


def _row_to_dict(row, columns, table, expand_times):
    data = {}
    for name in columns:
        value = getattr(row, name, None)
        data[name] = jsonify_value(value)
        if expand_times and hasattr(value, "isoformat") and hasattr(value, "hour"):
            from core.config import semantic_for

            data[f"{name}_utc"] = iso(as_utc(value, semantic_for(table, name)))
    return data


def _iter_rows(model, table, after_id=None, since_id=None, limit=None):
    """Stream rows in id order using a chunked server-side cursor."""
    query = model.query.order_by(model.id.asc())
    start = after_id if after_id is not None else since_id
    if start is not None:
        query = query.filter(model.id > start)
    if limit:
        query = query.limit(limit)
    return query.yield_per(CHUNK_SIZE)


# ---------------------------------------------------------------------------
# Manifest
# ---------------------------------------------------------------------------


@bp.get("/dump/manifest")
@require_auth("dump:read")
def manifest():
    """
    What is available to dump: tables, columns, row counts, id ranges.

    Fetch this first in a cron job - it tells you what to expect, and the
    ``max_id`` values become the ``since`` cursor for the next incremental run.
    """
    from extensions import db

    names = _requested_tables()
    tables = []

    for name in names:
        model = _resolve(name)
        try:
            count = db.session.query(db.func.count(model.id)).scalar() or 0
            bounds = db.session.query(
                db.func.min(model.id), db.func.max(model.id)
            ).one()
        except Exception:
            db.session.rollback()
            tables.append({"table": name, "error": "unreadable"})
            continue

        tables.append(
            {
                "table": name,
                "rows": int(count),
                "min_id": bounds[0],
                "max_id": bounds[1],
                "columns": _columns(model, name),
                "redacted_columns": sorted(REDACTED_COLUMNS.get(name, set())),
                "endpoint": f"/api/v1/dump/{name}",
            }
        )

    return ok(
        {
            "generated_at_utc": iso(as_utc(utc_now(), UTC_WALL)),
            "tables": tables,
            "total_rows": sum(t.get("rows", 0) for t in tables),
            "formats": list(FORMATS),
            "full_dump_endpoint": "/api/v1/dump",
            "notes": {
                "streaming": "ndjson streams; use it for anything large.",
                "incremental": "Pass ?since=<max_id from a previous manifest> per table.",
                "resume": "Pass ?after_id=<last id received> to continue an interrupted dump.",
                "verify": "GET /api/v1/dump/{table}/checksum to confirm integrity.",
            },
        }
    )


@bp.get("/dump/tables")
@require_auth("dump:read")
def list_tables():
    return ok(
        {
            "portal_tables": sorted(PORTAL_TABLES),
            "api_tables": sorted(API_TABLES),
            "default": sorted(PORTAL_TABLES),
            "audit_tables_require_scope": ["api_login_events", "api_request_logs"],
        }
    )


@bp.get("/dump/<table>/checksum")
@require_auth("dump:read")
def checksum(table):
    """
    Content checksum for a table, so the receiver can verify what it stored.

    SHA-256 over the same NDJSON bytes the dump endpoint would emit.
    """
    model = _resolve(table)
    columns = _columns(model, table)

    digest = hashlib.sha256()
    rows = 0
    for row in _iter_rows(model, table):
        line = json.dumps(
            _row_to_dict(row, columns, table, False), sort_keys=True, separators=(",", ":")
        )
        digest.update(line.encode())
        digest.update(b"\n")
        rows += 1

    return ok(
        {
            "table": table,
            "rows": rows,
            "algorithm": "sha256",
            "checksum": digest.hexdigest(),
            "note": "Computed over canonical NDJSON (sorted keys, no expanded times).",
        }
    )


# ---------------------------------------------------------------------------
# Serialisation formats
# ---------------------------------------------------------------------------


def _ndjson_stream(model, table, columns, expand_times, after_id, since_id, limit):
    for row in _iter_rows(model, table, after_id, since_id, limit):
        yield json.dumps(_row_to_dict(row, columns, table, expand_times), default=str) + "\n"


def _json_stream(model, table, columns, expand_times, after_id, since_id, limit):
    """Buffered-looking JSON, still produced incrementally."""
    yield '{"table":' + json.dumps(table) + ',"rows":['
    first = True
    for row in _iter_rows(model, table, after_id, since_id, limit):
        if not first:
            yield ","
        first = False
        yield json.dumps(_row_to_dict(row, columns, table, expand_times), default=str)
    yield "]}"


def _csv_stream(model, table, columns, expand_times, after_id, since_id, limit):
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")

    writer.writerow(columns)
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate(0)

    for row in _iter_rows(model, table, after_id, since_id, limit):
        data = _row_to_dict(row, columns, table, False)
        writer.writerow([data.get(name) for name in columns])
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)
    _ = expand_times


def _sql_literal(value):
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def _sql_stream(model, table, columns, expand_times, after_id, since_id, limit):
    column_list = ", ".join(f'"{name}"' for name in columns)
    yield f"-- VIAMA dump of {table}\n"
    yield f"-- generated {iso(as_utc(utc_now(), UTC_WALL))}\n"
    for row in _iter_rows(model, table, after_id, since_id, limit):
        data = _row_to_dict(row, columns, table, False)
        values = ", ".join(_sql_literal(data.get(name)) for name in columns)
        yield f'INSERT INTO "{table}" ({column_list}) VALUES ({values});\n'
    _ = expand_times


STREAMERS = {
    "ndjson": (_ndjson_stream, "application/x-ndjson", "ndjson"),
    "json": (_json_stream, "application/json", "json"),
    "csv": (_csv_stream, "text/csv", "csv"),
    "sql": (_sql_stream, "application/sql", "sql"),
}


def _gzip_wrap(chunks):
    """Compress a text generator on the fly, without buffering the whole body."""
    buffer = io.BytesIO()
    compressor = gzip.GzipFile(fileobj=buffer, mode="wb")

    def take():
        buffer.seek(0)
        data = buffer.read()
        buffer.seek(0)
        buffer.truncate(0)
        return data

    for chunk in chunks:
        compressor.write(chunk.encode("utf-8") if isinstance(chunk, str) else chunk)
        payload = take()
        if payload:
            yield payload
    compressor.close()
    payload = take()
    if payload:
        yield payload


def _stream_response(generator, mimetype, filename, compress):
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-Accel-Buffering": "no",
        "Cache-Control": "no-store",
    }
    if compress:
        headers["Content-Encoding"] = "gzip"
        generator = _gzip_wrap(generator)
    return Response(
        stream_with_context(generator), mimetype=mimetype, headers=headers
    )


def _dump_args():
    fmt = (request.args.get("format") or "ndjson").lower()
    if fmt not in FORMATS:
        raise ValidationError(
            message=f"Unknown format '{fmt}'.",
            details=[{"field": "format", "issue": "unknown", "value": fmt,
                      "allowed": list(FORMATS)}],
        )
    # "gzip" is the natural thing to write, so accept it as well as a boolean.
    raw_compress = (request.args.get("compress") or "false").strip().lower()
    compress = raw_compress == "gzip" or parse_bool(raw_compress, "compress")

    return {
        "format": fmt,
        "compress": compress,
        "expand_times": parse_bool(
            request.args.get("expand_times", "false"), "expand_times"
        ),
        "after_id": parse_int(request.args["after_id"], "after_id")
        if request.args.get("after_id")
        else None,
        "since": parse_int(request.args["since"], "since")
        if request.args.get("since")
        else None,
        "limit": parse_int(request.args["limit"], "limit")
        if request.args.get("limit")
        else None,
    }


# ---------------------------------------------------------------------------
# Single table
# ---------------------------------------------------------------------------


@bp.get("/dump/<table>")
@require_auth("dump:read")
def dump_table(table):
    """
    Stream one table.

    ``?format=ndjson|json|csv|sql``  ``?compress=gzip``
    ``?since=<id>`` incremental   ``?after_id=<id>`` resume   ``?limit=<n>``
    """
    model = _resolve(table)
    if table in ("api_login_events", "api_request_logs"):
        from core.engine import current_client

        client = current_client()
        if client:
            client.require("audit:read")

    args = _dump_args()
    columns = _columns(model, table)
    streamer, mimetype, extension = STREAMERS[args["format"]]

    stamp = utc_now().strftime("%Y%m%d")
    filename = f"viama_{table}_{stamp}.{extension}" + (".gz" if args["compress"] else "")

    generator = streamer(
        model,
        table,
        columns,
        args["expand_times"],
        args["after_id"],
        args["since"],
        args["limit"],
    )
    return _stream_response(generator, mimetype, filename, args["compress"])


# ---------------------------------------------------------------------------
# Whole database
# ---------------------------------------------------------------------------


@bp.get("/dump")
@require_auth("dump:read")
def dump_all():
    """
    Stream the entire database as one NDJSON document.

    Each line is ``{"_table": "<name>", ...row}``, preceded by a ``_meta`` line
    and followed by a ``_summary`` line carrying per-table counts - so the
    receiver can verify completeness without a second request.

    This is the endpoint a nightly cron should call:

        curl -sS -H "Authorization: Bearer $TOKEN" \\
             "https://<portal>/api/v1/dump?compress=gzip" \\
             -o "viama_$(date +%F).ndjson.gz"
    """
    names = _requested_tables()
    args = _dump_args()

    if args["format"] != "ndjson":
        raise ValidationError(
            message="The full dump is NDJSON only. Use /dump/{table} for other formats.",
            details=[{"field": "format", "issue": "must be ndjson for the full dump"}],
        )

    since_map = {}
    raw_since = request.args.get("since_map")
    if raw_since:
        try:
            since_map = json.loads(raw_since)
        except ValueError:
            raise ValidationError(
                message="'since_map' must be JSON, e.g. {\"surveys\": 812}.",
                details=[{"field": "since_map", "issue": "invalid JSON"}],
            )

    def generate():
        started = utc_now()
        yield json.dumps(
            {
                "_meta": {
                    "generated_at_utc": iso(as_utc(started, UTC_WALL)),
                    "api_version": "v1",
                    "tables": names,
                    "incremental": bool(since_map),
                    "format": "ndjson",
                }
            },
            default=str,
        ) + "\n"

        counts = {}
        for name in names:
            model = _resolve(name)
            columns = _columns(model, name)
            written = 0
            last_id = None
            for row in _iter_rows(model, name, since_id=since_map.get(name)):
                payload = _row_to_dict(row, columns, name, args["expand_times"])
                payload["_table"] = name
                yield json.dumps(payload, default=str) + "\n"
                written += 1
                last_id = getattr(row, "id", None)
            counts[name] = {"rows": written, "max_id": last_id}

        finished = utc_now()
        yield json.dumps(
            {
                "_summary": {
                    "tables": counts,
                    "total_rows": sum(c["rows"] for c in counts.values()),
                    "started_at_utc": iso(as_utc(started, UTC_WALL)),
                    "finished_at_utc": iso(as_utc(finished, UTC_WALL)),
                    "duration_seconds": round((finished - started).total_seconds(), 2),
                    # Feed these back as ?since_map= next time for a delta pull.
                    "next_since_map": {
                        name: c["max_id"] for name, c in counts.items() if c["max_id"]
                    },
                }
            },
            default=str,
        ) + "\n"

    stamp = utc_now().strftime("%Y%m%d")
    filename = f"viama_full_{stamp}.ndjson" + (".gz" if args["compress"] else "")
    return _stream_response(
        generate(), "application/x-ndjson", filename, args["compress"]
    )


@bp.get("/dump/status")
@require_auth("dump:read")
def dump_status():
    """Quick pre-flight: row counts and a rough size estimate."""
    from extensions import db

    names = _requested_tables()
    tables = {}
    total = 0
    for name in names:
        model = _resolve(name)
        try:
            count = int(db.session.query(db.func.count(model.id)).scalar() or 0)
        except Exception:
            db.session.rollback()
            count = None
        column_count = len(_columns(model, name))
        tables[name] = {
            "rows": count,
            "columns": column_count,
            "estimated_bytes": (count or 0) * column_count * 24,
        }
        total += tables[name]["estimated_bytes"]

    return ok(
        {
            "tables": tables,
            "estimated_total_bytes": total,
            "estimated_total_mb": round(total / (1024 * 1024), 2),
            "recommendation": (
                "Use ?compress=gzip. Expect roughly 8-12x compression on this data."
            ),
        }
    )

# ==========================================================================
# res_sync
# /api/v1/sync - the catch-up channel.
#
# Webhooks are the fast path; this is the reliable one.  If a delivery is missed,
# the endpoint was down, or the receiver simply wants to reconcile, it replays the
# change log from a cursor.
#
# The cursor is ``change_log.id``, not a timestamp, and that choice matters:
#
#     Sequence values are allocated at INSERT but rows only become visible at
#     COMMIT.  A transaction that took id=100 can commit *after* one that took
#     id=101.  A poller that reads up to 101 and stores cursor=101 will never see
#     row 100 - silently, forever.
#
# The fix is a visibility lag: only return rows older than SAFE_LAG_SECONDS, by
# which time any in-flight transaction has committed or rolled back.  The feed is
# therefore always a few seconds behind, which is correct for a catch-up channel.
# ==========================================================================

import os

from flask import Blueprint, request

from core.engine import require_auth
from core.engine import ok
from core.engine import ValidationError
from core.engine import parse_int
from core.engine import jsonify_value
from core.config import UTC_WALL, as_ist, as_utc, iso, utc_now


DEFAULT_LIMIT = 200
MAX_LIMIT = 500


def safe_lag_seconds():
    try:
        return int(os.getenv("SYNC_SAFE_LAG_SECONDS", "5"))
    except ValueError:
        return 5


def _change_row(row):
    return {
        "cursor": row.id,
        "entity": row.entity,
        "entity_id": row.entity_id,
        "event": row.event,
        "op": row.op,
        "actor": row.actor,
        "source": row.source,
        "data": jsonify_value(row.data),
        "previous": jsonify_value(row.previous),
        "created_at_utc": iso(as_utc(row.created_at, UTC_WALL)),
        "created_at_ist": iso(as_ist(row.created_at, UTC_WALL)),
    }


@bp.get("/sync/changes")
@require_auth("surveys:read")
def changes():
    """
    Every change since a cursor.

    ``?since=<cursor>`` - the ``next_cursor`` from your last call (start at 0).
    ``?entity=survey,assignment`` and ``?event=survey.completed`` narrow it.
    ``?since_time=<ISO>`` resolves a timestamp to a cursor for a first run.

    Poll this on a timer and you have a complete, ordered, gap-free replica -
    independent of whether any webhook was delivered.
    """
    from datetime import timedelta

    from core.models import ChangeLog

    since = parse_int(request.args.get("since", 0), "since")
    limit = min(parse_int(request.args.get("limit", DEFAULT_LIMIT), "limit"), MAX_LIMIT)

    lag = safe_lag_seconds()
    cutoff = utc_now() - timedelta(seconds=lag)

    query = ChangeLog.query.filter(ChangeLog.id > since, ChangeLog.created_at <= cutoff)

    entities = request.args.get("entity")
    if entities:
        query = query.filter(
            ChangeLog.entity.in_([e.strip() for e in entities.split(",") if e.strip()])
        )

    events = request.args.get("event")
    if events:
        query = query.filter(
            ChangeLog.event.in_([e.strip() for e in events.split(",") if e.strip()])
        )

    since_time = request.args.get("since_time")
    if since_time:
        from core.config import TimeParseError, parse_datetime

        try:
            instant = parse_datetime(since_time, UTC_WALL, assume_tz="UTC")
        except TimeParseError as exc:
            raise ValidationError(
                message=str(exc),
                details=[{"field": "since_time", "issue": "invalid datetime"}],
            )
        query = query.filter(ChangeLog.created_at >= instant)

    rows = query.order_by(ChangeLog.id.asc()).limit(limit + 1).all()
    has_more = len(rows) > limit
    rows = rows[:limit]

    return ok(
        {
            "changes": [_change_row(row) for row in rows],
            "next_cursor": rows[-1].id if rows else since,
            "has_more": has_more,
            "count": len(rows),
            "safe_lag_seconds": lag,
            "note": (
                f"Rows newer than {lag}s are withheld until their transaction is "
                "certain to have committed. Without this the feed would silently "
                "skip rows whose sequence value was allocated before, but "
                "committed after, one you already read."
            ),
        }
    )


@bp.get("/sync/cursor")
@require_auth("surveys:read")
def current_cursor():
    """
    The newest safe cursor.

    Call this once before a full backfill, then use it as ``since`` afterwards -
    that way nothing that happens during the backfill is lost.
    """
    from datetime import timedelta

    from extensions import db

    from core.models import ChangeLog

    lag = safe_lag_seconds()
    cutoff = utc_now() - timedelta(seconds=lag)

    latest = (
        db.session.query(db.func.max(ChangeLog.id))
        .filter(ChangeLog.created_at <= cutoff)
        .scalar()
    )
    absolute = db.session.query(db.func.max(ChangeLog.id)).scalar()

    return ok(
        {
            "cursor": latest or 0,
            "absolute_latest": absolute or 0,
            "withheld": (absolute or 0) - (latest or 0),
            "safe_lag_seconds": lag,
            "server_time_utc": iso(as_utc(utc_now(), UTC_WALL)),
        }
    )


@bp.get("/sync/status")
@require_auth("surveys:read")
def sync_status():
    """How far behind a given cursor is - use it to alert on replica lag."""
    from extensions import db

    from core.models import ChangeLog, WebhookOutbox

    since = parse_int(request.args.get("since", 0), "since")
    behind = ChangeLog.query.filter(ChangeLog.id > since).count()
    oldest = (
        ChangeLog.query.filter(ChangeLog.id > since)
        .order_by(ChangeLog.id.asc())
        .first()
    )

    return ok(
        {
            "since": since,
            "changes_behind": behind,
            "oldest_unsynced_utc": iso(as_utc(oldest.created_at, UTC_WALL))
            if oldest
            else None,
            "webhook_backlog": WebhookOutbox.query.filter_by(status="pending").count(),
            "total_changes": db.session.query(db.func.count(ChangeLog.id)).scalar() or 0,
        }
    )


@bp.get("/sync/entities")
@require_auth("surveys:read")
def entities():
    """Which entities appear in the feed, and how much of each."""
    from extensions import db

    from core.models import ChangeLog

    rows = (
        db.session.query(
            ChangeLog.entity,
            db.func.count(ChangeLog.id),
            db.func.max(ChangeLog.id),
        )
        .group_by(ChangeLog.entity)
        .all()
    )
    return ok(
        {
            "entities": [
                {"entity": entity, "changes": int(count), "latest_cursor": int(latest)}
                for entity, count, latest in rows
            ]
        }
    )

# ==========================================================================
# res_webhooks
# /api/v1/webhooks - manage push subscriptions and inspect deliveries.
# ==========================================================================

from flask import Blueprint, request

from core.engine import require_auth
from core.engine import commit, json_body
from core.engine import created, no_content, ok, paginated
from core.engine import NotFound, ValidationError
from core.engine import EVENT_CATALOG, enabled_events
from core.engine import ListParams, ResourceConfig, parse_int
from core.engine import jsonify_value
from core.config import UTC_WALL, as_utc, iso
from core.engine import drain_outbox, send_test



def _hook_row(row, hide=("secret",)):
    data = {}
    for column in row.__table__.columns:
        if column.key in hide:
            continue
        value = getattr(row, column.key)
        if column.key.endswith("_at") or column.key == "created_at":
            data[column.key + "_utc"] = iso(as_utc(value, UTC_WALL))
        else:
            data[column.key] = jsonify_value(value)
    return data


@bp.get("/webhooks/events")
@require_auth("webhooks:admin")
def catalog():
    """Every event that can be delivered, and which are currently enabled."""
    allow = enabled_events()
    return ok(
        {
            "events": [
                {
                    "event": name,
                    "description": description,
                    "enabled": (
                        True
                        if allow == "*"
                        else (name in allow if allow else name != "assignment.updated")
                    ),
                }
                for name, description in sorted(EVENT_CATALOG.items())
            ],
            "wildcards": ["*", "survey.*", "assignment.*", "user.*", "auth.*"],
            "muted_by_default": ["assignment.updated"],
            "note": (
                "assignment.updated is muted because the portal rewrites assignment "
                "rows during ordinary dashboard loads. Set WEBHOOK_ENABLED_EVENTS=* "
                "to receive everything."
            ),
            "signature": {
                "header": "X-Viama-Signature",
                "format": "t=<unix>,v1=<hex hmac_sha256(secret, '<t>.<raw body>')>",
                "tolerance_seconds": 300,
                "advice": "Run NTP on the receiver; clock skew is the usual cause of failures.",
            },
        }
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@bp.get("/webhooks/endpoints")
@require_auth("webhooks:admin")
def list_endpoints():
    from core.models import WebhookEndpoint

    rows = WebhookEndpoint.query.order_by(WebhookEndpoint.id).all()
    return ok([_hook_row(row) for row in rows])


@bp.post("/webhooks/endpoints")
@require_auth("webhooks:admin")
def create_endpoint():
    """
    Register a URL to receive events.

    The signing secret is returned ONCE. Store it - the receiver needs it to
    verify signatures, and it is never shown again.
    """
    import secrets

    from extensions import db

    from core.models import WebhookEndpoint

    body = json_body()
    url = (body.get("url") or "").strip()
    if not url.startswith(("http://", "https://")):
        raise ValidationError(
            message="'url' must be an absolute http(s) URL.",
            details=[{"field": "url", "issue": "must start with http:// or https://"}],
        )

    events = body.get("events", "*")
    if isinstance(events, list):
        events = ",".join(events)

    endpoint = WebhookEndpoint(
        url=url,
        secret=body.get("secret") or secrets.token_urlsafe(32),
        events=events,
        active=bool(body.get("active", True)),
        description=body.get("description"),
    )
    db.session.add(endpoint)
    commit("create endpoint")

    payload = _hook_row(endpoint, hide=())
    payload["_warning"] = "The secret is shown only on creation. Store it now."
    return created(payload, location=f"/api/v1/webhooks/endpoints/{endpoint.id}")


@bp.get("/webhooks/endpoints/<int:endpoint_id>")
@require_auth("webhooks:admin")
def show_endpoint(endpoint_id):
    return ok(_hook_row(_get_endpoint(endpoint_id)))


@bp.patch("/webhooks/endpoints/<int:endpoint_id>")
@require_auth("webhooks:admin")
def patch_endpoint(endpoint_id):
    endpoint = _get_endpoint(endpoint_id)
    body = json_body()

    if "url" in body:
        endpoint.url = body["url"]
    if "events" in body:
        endpoint.events = (
            ",".join(body["events"]) if isinstance(body["events"], list) else body["events"]
        )
    if "active" in body:
        endpoint.active = bool(body["active"])
    if "description" in body:
        endpoint.description = body["description"]
    if body.get("rotate_secret"):
        import secrets

        endpoint.secret = secrets.token_urlsafe(32)

    commit("update endpoint")
    payload = _hook_row(endpoint, hide=() if body.get("rotate_secret") else ("secret",))
    return ok(payload)


@bp.delete("/webhooks/endpoints/<int:endpoint_id>")
@require_auth("webhooks:admin")
def delete_endpoint(endpoint_id):
    from extensions import db

    db.session.delete(_get_endpoint(endpoint_id))
    commit("delete endpoint")
    return no_content()


@bp.post("/webhooks/endpoints/<int:endpoint_id>/test")
@require_auth("webhooks:admin")
def test_endpoint(endpoint_id):
    """Send a signed ``ping`` so you can prove the receiver works."""
    body = json_body(required=False)
    return ok(send_test(_get_endpoint(endpoint_id), event=body.get("event", "ping")))


def _get_endpoint(endpoint_id):
    from core.models import WebhookEndpoint

    endpoint = WebhookEndpoint.query.get(endpoint_id)
    if not endpoint:
        raise NotFound(message=f"Webhook endpoint {endpoint_id} not found.")
    return endpoint


# ---------------------------------------------------------------------------
# Outbox / deliveries
# ---------------------------------------------------------------------------


@bp.post("/webhooks/drain")
@require_auth("webhooks:admin")
def drain():
    """
    Deliver pending events now.

    This is the endpoint to call on a schedule - from Vercel Cron, or from the
    consuming VM at whatever interval it wants. Delivery latency is the drain
    interval, so the consumer controls it.
    """
    body = json_body(required=False)
    limit = min(int(body.get("max", 50)), 200)
    return ok(drain_outbox(limit=limit, endpoint_id=body.get("endpoint_id")))


@bp.get("/webhooks/outbox")
@require_auth("webhooks:admin")
def list_outbox():
    from core.models import WebhookOutbox

    config = ResourceConfig(
        model=WebhookOutbox,
        table="webhook_outbox",
        spec_name="outbox",
        sortable=("id", "created_at", "event", "status", "attempts"),
        default_sort="-id",
    )
    params = ListParams.from_request(config)
    query = params.apply(WebhookOutbox.query)
    rows, page_info = params.paginate(query)
    return paginated([_hook_row(row, hide=()) for row in rows], page_info, meta=params.meta())


@bp.get("/webhooks/deliveries")
@require_auth("webhooks:admin")
def list_deliveries():
    from core.models import WebhookDelivery

    config = ResourceConfig(
        model=WebhookDelivery,
        table="webhook_deliveries",
        spec_name="delivery",
        sortable=("id", "created_at", "event", "status", "endpoint_id", "latency_ms"),
        default_sort="-id",
    )
    params = ListParams.from_request(config)
    query = params.apply(WebhookDelivery.query)
    rows, page_info = params.paginate(query)
    return paginated([_hook_row(row, hide=()) for row in rows], page_info, meta=params.meta())


@bp.post("/webhooks/outbox/<int:outbox_id>/retry")
@require_auth("webhooks:admin")
def retry_outbox(outbox_id):
    """Re-queue a dead-lettered event and attempt it immediately."""
    from core.models import WebhookOutbox

    row = WebhookOutbox.query.get(outbox_id)
    if not row:
        raise NotFound(message=f"Outbox entry {outbox_id} not found.")

    row.status = "pending"
    row.next_retry_at = None
    row.attempts = 0
    commit("retry")

    return ok(drain_outbox(limit=1))


@bp.get("/webhooks/health")
@require_auth("webhooks:admin")
def webhook_health():
    """Backlog and failure summary - worth alerting on."""
    from extensions import db

    from core.models import WebhookEndpoint, WebhookOutbox

    by_status = dict(
        db.session.query(WebhookOutbox.status, db.func.count(WebhookOutbox.id))
        .group_by(WebhookOutbox.status)
        .all()
    )
    oldest = (
        WebhookOutbox.query.filter_by(status="pending")
        .order_by(WebhookOutbox.id.asc())
        .first()
    )
    endpoints = WebhookEndpoint.query.all()

    return ok(
        {
            "outbox": {k: int(v) for k, v in by_status.items()},
            "oldest_pending_utc": iso(as_utc(oldest.created_at, UTC_WALL)) if oldest else None,
            "endpoints": [
                {
                    "id": e.id,
                    "url": e.url,
                    "active": e.active,
                    "failure_count": e.failure_count,
                    "last_success_utc": iso(as_utc(e.last_success_at, UTC_WALL)),
                    "last_error": e.last_error,
                }
                for e in endpoints
            ],
            "healthy": not any(e.failure_count and e.failure_count > 3 for e in endpoints),
        }
    )


_ = (parse_int, request)

# ==========================================================================
# res_jobs
# /api/v1/jobs - scheduled maintenance.
#
# ``POST /jobs/tick`` is a single entrypoint that decides internally what is due:
# drain webhooks, run the missed-survey engine, do the Monday reset, purge expired
# audit rows and idempotency keys.
#
# One endpoint rather than five because Vercel's Hobby plan allows only two cron
# jobs at daily granularity - consolidating keeps the whole thing inside one slot
# and works on any plan.  If you need finer cadence than your plan allows, have the
# consuming VM call this instead; it is the same work either way.
#
# Auth: ``X-Cron-Secret`` matching ``CRON_SECRET`` (Vercel also sends it as a bearer
# token when that env var is set), or any API token holding ``jobs:run``.
# ==========================================================================

from flask import Blueprint, request

from core.engine import require_auth, require_cron_secret
from core.engine import json_body
from core.engine import ok
from core.config import UTC_WALL, as_utc, ist_now, iso, utc_now



@bp.post("/jobs/tick")
def tick():
    """
    Run everything that is due. Idempotent and safe to call often.

    ``?only=webhooks,alerts`` restricts which jobs run; ``?dry_run=true``
    reports without writing.
    """
    require_cron_secret()

    body = json_body(required=False)
    only = request.args.get("only") or body.get("only")
    wanted = (
        {job.strip() for job in only.split(",") if job.strip()}
        if isinstance(only, str)
        else set(only or [])
    )
    dry_run = str(request.args.get("dry_run", body.get("dry_run", ""))).lower() in (
        "1",
        "true",
        "yes",
    )

    def should(name):
        return not wanted or name in wanted

    started = utc_now()
    results = {}

    if should("webhooks"):
        results["webhooks"] = _safely(lambda: _drain(dry_run))

    if should("alerts"):
        results["missed_engine"] = _safely(lambda: _alerts(dry_run))

    if should("weekly_reset"):
        results["weekly_reset"] = _safely(lambda: _weekly_reset(dry_run))

    if should("purge"):
        results["purge"] = _safely(lambda: _purge(dry_run))

    finished = utc_now()
    return ok(
        {
            "ran_at_utc": iso(as_utc(started, UTC_WALL)),
            "ran_at_ist": ist_now().isoformat(),
            "duration_seconds": round((finished - started).total_seconds(), 2),
            "dry_run": dry_run,
            "jobs": results,
        }
    )


def _safely(func):
    """One failing job must not stop the others."""
    try:
        return func()
    except Exception as exc:
        return {"error": f"{type(exc).__name__}: {exc}"[:300]}


def _drain(dry_run):
    from core.models import WebhookOutbox
    from core.engine import drain_outbox

    if dry_run:
        return {
            "skipped": "dry_run",
            "pending": WebhookOutbox.query.filter_by(status="pending").count(),
        }
    return drain_outbox(limit=100)


def _alerts(dry_run):
    from core.engine import run_missed_engine

    now = ist_now()
    from core.config import ALERT_CUTOFF_HOUR_PRIMARY

    if now.hour < ALERT_CUTOFF_HOUR_PRIMARY:
        return {
            "skipped": f"before {ALERT_CUTOFF_HOUR_PRIMARY}:00 IST",
            "hour_ist": now.hour,
        }
    return run_missed_engine(dry_run=dry_run)


def _weekly_reset(dry_run):
    from core.engine import run_weekly_reset

    return run_weekly_reset(dry_run=dry_run)


def _purge(dry_run):
    from datetime import timedelta

    from sqlalchemy.orm import Session

    from extensions import db

    from core.models import IdempotencyKey, WebhookDelivery
    from core.engine import purge_old, retention_days

    if dry_run:
        return {"skipped": "dry_run", "retention_days": retention_days()}

    result = purge_old()

    # Idempotency claims are only useful for the retry window.
    session = Session(bind=db.engine)
    try:
        cutoff = utc_now() - timedelta(hours=24)
        result["idempotency_keys_removed"] = int(
            session.query(IdempotencyKey)
            .filter(IdempotencyKey.created_at < cutoff)
            .delete(synchronize_session=False)
            or 0
        )
        delivery_cutoff = utc_now() - timedelta(days=30)
        result["webhook_deliveries_removed"] = int(
            session.query(WebhookDelivery)
            .filter(WebhookDelivery.created_at < delivery_cutoff)
            .delete(synchronize_session=False)
            or 0
        )
        session.commit()
    except Exception:
        session.rollback()
        result["purge_error"] = "partial"
    finally:
        session.close()

    return result


@bp.get("/jobs")
@require_auth("jobs:run")
def list_jobs():
    """What /jobs/tick will do, and how to schedule it."""
    from core.engine import retention_days
    from core.config import ALERT_CUTOFF_HOUR_PRIMARY

    return ok(
        {
            "jobs": [
                {
                    "name": "webhooks",
                    "does": "Delivers pending outbox events to subscribed endpoints.",
                    "frequency": "as often as you want push latency to be",
                },
                {
                    "name": "alerts",
                    "does": "Runs the missed-survey engine and persists status changes.",
                    "frequency": f"a few times a day after {ALERT_CUTOFF_HOUR_PRIMARY}:00 IST",
                },
                {
                    "name": "weekly_reset",
                    "does": "Resets all assignments to 'assigned' on a Monday.",
                    "frequency": "daily; it no-ops on other days",
                },
                {
                    "name": "purge",
                    "does": (
                        f"Deletes audit rows older than {retention_days()} days, "
                        "idempotency keys older than 24h, deliveries older than 30d."
                    ),
                    "frequency": "daily",
                },
            ],
            "endpoint": "POST /api/v1/jobs/tick",
            "auth": "X-Cron-Secret: <CRON_SECRET>, or a token with jobs:run",
            "vercel_cron": {
                "note": "Add to vercel.json. Hobby allows 2 crons at daily granularity.",
                "crons": [{"path": "/api/v1/jobs/tick", "schedule": "*/15 * * * *"}],
            },
            "external_cron": (
                "*/15 * * * * curl -fsS -X POST "
                "-H 'X-Cron-Secret: $CRON_SECRET' "
                "https://<portal>/api/v1/jobs/tick"
            ),
        }
    )

# ==========================================================================
# res_dashboards
# /api/v1/dashboards - one endpoint per portal screen.
#
# Each returns the same data the corresponding page renders, under the same keys,
# so the other site can rebuild any screen 1:1 without re-deriving anything.
#
# Two things are deliberately preserved rather than "fixed":
#
# * ``display_start_time`` / ``display_end_time`` reproduce the portal's on-screen
#   value, which for ``start_time`` is ``raw + 5:30`` applied to a column that is
#   already IST - i.e. 5h30m ahead of reality.  Use ``start_time_utc`` for the
#   truth; use ``display_*`` when you want the number the portal shows.
#
# * Fields the pages hardcode as empty stay empty: the regional dashboard passes
#   ``alerts=[]`` and ``resurvey_requests=[]`` (routes/regional.py:230,359) and the
#   team-leader dashboard passes ``missed=0``.  Set ``?live=true`` to compute the
#   real values instead.
#
# Unlike the portal, none of these endpoints write anything.
# ==========================================================================

from flask import Blueprint, request

from core.engine import require_auth, resolve_actor
from core.config import (
    SURVEY_COMPLETED,
    SURVEY_GROUNDWORK_COMPLETED,
    SURVEY_ONGOING,
    SURVEY_VIDEO_PENDING,
)
from core.engine import base_query, get_or_404, wants_deleted
from core.config import (
    attach_derived,
    day_counts,
    day_counts_global,
    defect_counts,
    team_km_totals,
)
from core.engine import ok
from core.engine import Forbidden
from core.engine import ListParams
from core.engine import Options, serialize_many, spec, survey_derived
from core.engine import compute_alerts, missed_count
from core.config import week_list


DERIVED_ROW_FIELDS = (
    "scheduled_day",
    "display_start_time",
    "display_end_time",
    "upload_duration_minutes",
    "upload_status_text",
    "survey_duration_minutes",
)


def _flag(name, default=False):
    value = request.args.get(name)
    if value is None:
        return default
    return value.strip().lower() in ("1", "true", "yes")


def _rows(surveys):
    """Serialize surveys with the derived attributes flattened onto each row."""
    from core.config import iso_naive

    attach_derived(surveys)
    options = Options.from_request()
    items = serialize_many(surveys, spec("survey"), options)

    for item, survey in zip(items, surveys):
        for field in DERIVED_ROW_FIELDS:
            value = getattr(survey, field, None)
            item[field] = iso_naive(value) if hasattr(value, "isoformat") else value
    return items


def _survey_query(extra_allowed=()):
    from models.db_models import Survey

    from core.endpoints import survey_config

    cfg = survey_config()
    params = ListParams.from_request(cfg, extra_allowed=set(extra_allowed) | {"live"})
    query = params.apply_filters(base_query(Survey, wants_deleted()))
    return cfg, params, query


def _kpis(states=None):
    """The four stat cards. Counts are global, matching the pages."""
    from models.db_models import Survey, User

    from core.engine import base_query as scoped

    def count(status):
        query = scoped(Survey).filter(Survey.status == status)
        if states:
            query = query.filter(Survey.state.in_(states))
        return query.count()

    captains = User.query.filter_by(role="captain")
    if states:
        # regional.py:59 scopes captains by User.region, not User.state.
        captains = captains.filter(User.region.in_(states))

    return {
        "total_captains": captains.count(),
        "ongoing": count(SURVEY_ONGOING),
        "groundwork_completed": count(SURVEY_GROUNDWORK_COMPLETED),
        "completed": count(SURVEY_COMPLETED),
        "video_pending": count(SURVEY_VIDEO_PENDING),
        "missed": missed_count(states),
    }


def _dropdowns(states=None):
    from extensions import db
    from models.db_models import Survey, SurveyAssignment, User

    if states is None:
        state_rows = (
            db.session.query(SurveyAssignment.state)
            .distinct()
            .order_by(SurveyAssignment.state)
            .all()
        )
        state_values = [row[0] for row in state_rows if row[0]]
    else:
        state_values = list(states)

    captains = User.query.filter_by(role="captain")
    if states:
        captains = captains.filter(User.region.in_(states))

    cycles = db.session.query(Survey.cycle_no).distinct().order_by(Survey.cycle_no).all()

    return {
        "states": state_values,
        "captains": [
            {"id": c.id, "name": c.name, "email": c.email} for c in captains.order_by(User.name)
        ],
        "cycles": [row[0] for row in cycles if row[0] is not None],
        "weeks": week_list(),
    }


# ---------------------------------------------------------------------------
# Admin
# ---------------------------------------------------------------------------


@bp.get("/dashboards/admin")
@require_auth("dashboards:read")
def admin_dashboard():
    """
    GET /admin as JSON.

    Unlike the page, this never runs the missed-survey engine or the Monday
    reset - those are POST /alerts/evaluate and POST /assignments/weekly-reset.
    ``alerts`` here is computed read-only.
    """
    from models.db_models import Survey

    _cfg, params, query = _survey_query()
    query = query.filter(Survey.show_on_dashboard.isnot(False))

    if not request.args.get("sort"):
        from core.config import admin_rank_case

        query = query.order_by(admin_rank_case(), Survey.start_time.desc(), Survey.id.asc())
    else:
        query = params.apply_sort(query)

    surveys, page_info = params.paginate(query)
    all_surveys = _rows(surveys)

    # Team KM is computed over ALL completed surveys matching the filters, not
    # just the current page - the page has no pagination.
    km_source = query.limit(None).offset(None).all() if page_info.get("total") else surveys
    totals = team_km_totals(km_source)

    from core.endpoints import pending_resurveys  # noqa: F401  (documented sibling)

    resurveys = (
        base_query(Survey)
        .filter(Survey.resurvey_requested.is_(True), Survey.resurvey_approved.isnot(True))
        .order_by(Survey.id.desc())
        .all()
    )

    data = {
        **_kpis(),
        "krish_km": totals["Krish"],
        "godbole_km": totals["Godbole"],
        "aspizo_km": totals["Aspizo"],
        "total_km": totals["total"],
        "all_surveys": all_surveys,
        "resurvey_requests": serialize_many(resurveys, spec("survey"), Options.from_request()),
        "alerts": compute_alerts(variant="admin" if _flag("legacy_alerts") else "regional"),
        **_dropdowns(),
        "pagination": page_info,
    }
    return ok(data, meta=params.meta())


@bp.get("/dashboards/admin/missed")
@require_auth("dashboards:read")
def admin_missed():
    """GET /admin/missed - routes/admin.py:573-578."""
    from models.db_models import SurveyAssignment

    rows = (
        base_query(SurveyAssignment)
        .filter(SurveyAssignment.status == "missed")
        .order_by(SurveyAssignment.survey_day, SurveyAssignment.section_no)
        .all()
    )
    return ok(
        {
            "missed_surveys": serialize_many(rows, spec("assignment"), Options.from_request())
        }
    )


@bp.get("/dashboards/admin/schedules")
@require_auth("dashboards:read")
def admin_schedules():
    """
    GET /admin/schedules.

    The Mon-Fri cards are unfiltered COUNT queries on the page, so they ignore
    the active day/state filter. Reproduced; pass ``?respect_filters=true`` for
    counts that follow the filter instead.
    """
    from models.db_models import SurveyAssignment

    from core.config import day_order_case
    from core.endpoints import assignment_config

    cfg = assignment_config()
    params = ListParams.from_request(cfg, extra_allowed={"respect_filters", "live"})
    query = params.apply_filters(base_query(SurveyAssignment, wants_deleted()))
    query = query.order_by(
        day_order_case(), SurveyAssignment.state, SurveyAssignment.section_no
    )
    rows = query.all()

    counts = (
        day_counts(rows) if _flag("respect_filters") else day_counts_global()
    )

    return ok(
        {
            "schedules": serialize_many(rows, spec("assignment"), Options.from_request()),
            **counts,
            **_dropdowns(),
        },
        meta=params.meta(),
    )


@bp.get("/dashboards/admin/surveys/<int:survey_id>")
@require_auth("dashboards:read")
def admin_survey_details(survey_id):
    """GET /admin/survey/<id> - counts, dashcam photo, PDF, re-upload state."""
    from models.db_models import Survey

    survey = get_or_404(Survey, survey_id, wants_deleted())
    return ok(
        {
            "survey": spec("survey").serialize(survey, Options.from_request()),
            "derived": survey_derived(survey),
            "defect_counts": defect_counts(survey),
            "media": {
                "dashcam_photo": survey.dashcam_photo,
                "end_survey_pdf": survey.end_survey_pdf,
                "end_survey_photo": survey.end_survey_photo,
            },
            "pdf_reupload": {
                "required": bool(survey.pdf_reupload_required),
                "reason": survey.pdf_reupload_reason,
                "count": survey.pdf_reupload_count or 0,
            },
        }
    )


@bp.get("/dashboards/admin/reports")
@require_auth("dashboards:read")
def admin_reports():
    """GET /reports - the filtered table plus its four summary cards."""
    from models.db_models import Survey

    from core.config import report_summary

    _cfg, params, query = _survey_query()
    if not request.args.get("status"):
        query = query.filter(Survey.status == SURVEY_COMPLETED)

    surveys = query.all()
    return ok(
        {
            **report_summary(surveys),
            "surveys": serialize_many(surveys, spec("survey"), Options.from_request()),
            **_dropdowns(),
        },
        meta=params.meta(),
    )


@bp.get("/dashboards/admin/remark/<int:survey_id>")
@require_auth("dashboards:read")
def admin_remark(survey_id):
    from models.db_models import Survey

    from core.config import UTC_WALL, as_utc, iso

    survey = get_or_404(Survey, survey_id, wants_deleted())
    return ok(
        {
            "survey_id": survey.id,
            "section_no": survey.section_no,
            "upc_code": survey.upc_code,
            "captain_name": survey.captain_name,
            "roadvision_remark": survey.roadvision_remark,
            "roadvision_completed": bool(survey.roadvision_completed),
            "roadvision_completed_at_utc": iso(
                as_utc(survey.roadvision_completed_at, UTC_WALL)
            ),
        }
    )


@bp.get("/dashboards/admin/gmail-drafts")
@require_auth("dashboards:read")
def gmail_draft_options():
    """The week and section pickers on /admin/gmail-drafts/<type>."""
    from models.db_models import Survey

    from core.config import EMAIL_DRAFT_STATUSES, EMAIL_TYPES

    week = request.args.get("week")
    surveys = []
    if week:
        from core.config import week_window

        start, end = week_window(int(week))
        surveys = (
            base_query(Survey)
            .filter(
                Survey.status.in_(list(EMAIL_DRAFT_STATUSES)),
                Survey.start_time >= start,
                Survey.start_time < end,
            )
            .order_by(Survey.section_no)
            .all()
        )

    return ok(
        {
            "weeks": week_list(),
            "email_types": list(EMAIL_TYPES),
            "surveys": [
                {
                    "id": s.id,
                    "section_no": s.section_no,
                    "upc_code": s.upc_code,
                    "stretch_code": s.stretch_code,
                    "cycle_no": s.cycle_no,
                    "nh_number": s.nh_number,
                    "ro": s.ro,
                    "piu": s.piu,
                }
                for s in surveys
            ],
        }
    )


@bp.post("/dashboards/admin/gmail-drafts/<email_type>")
@require_auth("dashboards:read")
def build_gmail_draft(email_type):
    """
    Build the subject and HTML body for a survey email.

    Calls utils/email_templates.py unchanged, so the output matches the portal's
    exactly - including the ``{upc_code}_{cycle_no:03d}_{ddmmyy}`` survey ref.
    """
    from models.db_models import Survey

    from core.config import EMAIL_TYPES
    from core.engine import json_body
    from core.engine import ValidationError

    if email_type not in EMAIL_TYPES:
        raise ValidationError(
            message=f"Unknown email type '{email_type}'.",
            details=[{"field": "email_type", "issue": "unknown", "value": email_type,
                      "allowed": list(EMAIL_TYPES)}],
        )

    body = json_body()
    survey_id = body.get("survey_id")
    if not survey_id:
        raise ValidationError(
            message="'survey_id' is required.",
            details=[{"field": "survey_id", "issue": "required"}],
        )

    survey = get_or_404(Survey, survey_id, wants_deleted())
    start_date = body.get("start_date", "")
    end_date = body.get("end_date", "")

    from utils.email_templates import build_email_body, build_subject

    return ok(
        {
            "email_type": email_type,
            "subject": build_subject(survey, email_type, end_date),
            "body_html": build_email_body(survey, email_type, start_date, end_date),
            "survey": spec("survey").serialize(survey, Options.from_request()),
        }
    )


# ---------------------------------------------------------------------------
# Regional
# ---------------------------------------------------------------------------


def _manager_states(manager_email):
    from models.db_models import RegionalManagerState

    rows = RegionalManagerState.query.filter_by(manager_email=manager_email).all()
    return [row.state for row in rows]


@bp.get("/dashboards/regional")
@require_auth("dashboards:read")
def regional_dashboard():
    """
    GET /regional, scoped to the manager's states.

    ``alerts`` and ``resurvey_requests`` are ``[]`` because the page hardcodes
    them so (routes/regional.py:230,359). ``?live=true`` computes them for real.
    """
    from models.db_models import Survey

    manager = resolve_actor(roles=("regional_manager", "admin"))
    states = _manager_states(manager.email)

    _cfg, params, query = _survey_query()
    query = query.filter(
        Survey.state.in_(states or [""]), Survey.show_on_dashboard.isnot(False)
    )

    if not request.args.get("sort"):
        from core.config import regional_rank_case

        query = query.order_by(
            regional_rank_case(), Survey.start_time.desc(), Survey.id.asc()
        )
    else:
        query = params.apply_sort(query)

    surveys, page_info = params.paginate(query)
    live = _flag("live")

    kpis = _kpis(states)
    # regional.py:88 counts ongoing + groundwork together, unlike admin.
    kpis["ongoing"] = kpis["ongoing"] + kpis["groundwork_completed"]

    return ok(
        {
            **kpis,
            "manager_email": manager.email,
            "manager_states": states,
            "all_surveys": _rows(surveys),
            "alerts": compute_alerts(states=states) if live else [],
            "resurvey_requests": [],
            **_dropdowns(states),
            "pagination": page_info,
        },
        meta=params.meta(),
    )


@bp.get("/dashboards/regional/missed")
@require_auth("dashboards:read")
def regional_missed():
    from models.db_models import SurveyAssignment

    manager = resolve_actor(roles=("regional_manager", "admin"))
    states = _manager_states(manager.email)

    rows = (
        base_query(SurveyAssignment)
        .filter(
            SurveyAssignment.status == "missed",
            SurveyAssignment.state.in_(states or [""]),
        )
        .order_by(SurveyAssignment.survey_day, SurveyAssignment.section_no)
        .all()
    )
    return ok(
        {
            "manager_states": states,
            "missed_surveys": serialize_many(rows, spec("assignment"), Options.from_request()),
        }
    )


@bp.get("/dashboards/regional/schedules")
@require_auth("dashboards:read")
def regional_schedules():
    from models.db_models import SurveyAssignment

    manager = resolve_actor(roles=("regional_manager", "admin"))
    states = _manager_states(manager.email)

    rows = (
        base_query(SurveyAssignment)
        .filter(SurveyAssignment.state.in_(states or [""]))
        .order_by(SurveyAssignment.survey_day, SurveyAssignment.section_no)
        .all()
    )
    return ok(
        {
            "manager_states": states,
            "schedules": serialize_many(rows, spec("assignment"), Options.from_request()),
            **day_counts(rows),
        }
    )


@bp.get("/dashboards/regional/surveys/<int:survey_id>")
@require_auth("dashboards:read")
def regional_survey_details(survey_id):
    """403 outside the manager's states - mirrors routes/regional.py:458-459."""
    from models.db_models import Survey

    manager = resolve_actor(roles=("regional_manager", "admin"))
    states = _manager_states(manager.email)
    survey = get_or_404(Survey, survey_id, wants_deleted())

    if survey.state not in states and manager.role != "admin":
        raise Forbidden(
            message=f"Survey {survey_id} is in {survey.state}, outside your states.",
            code="state_not_authorized",
        )

    return ok(
        {
            "survey": spec("survey").serialize(survey, Options.from_request()),
            "derived": survey_derived(survey),
            "media": {
                "dashcam_photo": survey.dashcam_photo,
                # Rendered by the page but never written by any route.
                "end_survey_photo": survey.end_survey_photo,
            },
        }
    )


# ---------------------------------------------------------------------------
# Team leader
# ---------------------------------------------------------------------------


@bp.get("/dashboards/teamleader")
@require_auth("dashboards:read")
def teamleader_dashboard():
    """
    GET /teamleader.

    ``missed`` is 0 and the alert/resurvey scrollers are empty because the page
    hardcodes them. ``?live=true`` computes them.
    """
    from models.db_models import Survey

    _cfg, params, query = _survey_query()
    query = query.filter(Survey.show_in_teamleader_dashboard.isnot(False))

    if not request.args.get("sort"):
        from core.config import teamleader_rank_case

        query = query.order_by(
            teamleader_rank_case(), Survey.start_time.desc(), Survey.id.asc()
        )
    else:
        query = params.apply_sort(query)

    surveys, page_info = params.paginate(query)
    live = _flag("live")
    kpis = _kpis()
    if not live:
        kpis["missed"] = 0

    return ok(
        {
            **kpis,
            "all_surveys": _rows(surveys),
            "alerts": compute_alerts() if live else [],
            "resurvey_requests": [],
            "task_labels": {
                "survey_form": "Survey Form",
                "task1": "Raw Video",
                "task2": "Final Report",
            },
            **_dropdowns(),
            "pagination": page_info,
        },
        meta=params.meta(),
    )


@bp.get("/dashboards/teamleader/schedules")
@require_auth("dashboards:read")
def teamleader_schedules():
    from models.db_models import SurveyAssignment

    rows = (
        base_query(SurveyAssignment)
        .order_by(SurveyAssignment.survey_day, SurveyAssignment.section_no)
        .all()
    )
    return ok(
        {
            "schedules": serialize_many(rows, spec("assignment"), Options.from_request()),
            **day_counts(rows),
            **_dropdowns(),
        }
    )


@bp.get("/dashboards/teamleader/surveys/<int:survey_id>")
@require_auth("dashboards:read")
def teamleader_survey_details(survey_id):
    from models.db_models import Survey

    survey = get_or_404(Survey, survey_id, wants_deleted())
    return ok(
        {
            "survey": spec("survey").serialize(survey, Options.from_request()),
            "derived": survey_derived(survey),
            "media": {
                "dashcam_photo": survey.dashcam_photo,
                "end_survey_photo": survey.end_survey_photo,
            },
        }
    )


# ---------------------------------------------------------------------------
# RoadVision
# ---------------------------------------------------------------------------


@bp.get("/dashboards/roadvision")
@require_auth("dashboards:read")
def roadvision_dashboard():
    """GET /roadvision - completed surveys only, unreviewed first."""
    from models.db_models import Survey

    _cfg, params, query = _survey_query()
    query = query.filter(Survey.status == SURVEY_COMPLETED)

    if not request.args.get("sort"):
        from core.config import roadvision_rank_case

        query = query.order_by(
            roadvision_rank_case(), Survey.start_time.desc(), Survey.id.asc()
        )
    else:
        query = params.apply_sort(query)

    surveys, page_info = params.paginate(query)

    return ok(
        {
            **_kpis(),
            "all_surveys": _rows(surveys),
            "pending_review": base_query(Survey)
            .filter(
                Survey.status == SURVEY_COMPLETED,
                Survey.roadvision_completed.isnot(True),
            )
            .count(),
            **_dropdowns(),
            "pagination": page_info,
        },
        meta=params.meta(),
    )


@bp.get("/dashboards/roadvision/surveys/<int:survey_id>")
@require_auth("dashboards:read")
def roadvision_survey_details(survey_id):
    from models.db_models import Survey

    survey = get_or_404(Survey, survey_id, wants_deleted())
    return ok(
        {
            "survey": spec("survey").serialize(survey, Options.from_request()),
            "derived": survey_derived(survey),
            "defect_counts": defect_counts(survey),
            "video_count_matched": getattr(survey, "video_count_matched", None),
        }
    )


@bp.get("/dashboards/roadvision/remark/<int:survey_id>")
@require_auth("dashboards:read")
def roadvision_remark(survey_id):
    return admin_remark(survey_id)


# ---------------------------------------------------------------------------
# Captain (mobile)
# ---------------------------------------------------------------------------


def _captain_counts(captain, state_scoped=False):
    from models.db_models import Survey, SurveyAssignment

    if state_scoped:
        # Backup captains see everything in their state - routes/captain.py:961.
        assignments = base_query(SurveyAssignment).filter(
            SurveyAssignment.state == captain.state
        )
    else:
        assignments = base_query(SurveyAssignment).filter(
            SurveyAssignment.captain_email == captain.email
        )

    mine = base_query(Survey).filter(Survey.captain_email == captain.email)
    return {
        "assigned_count": assignments.count(),
        "pending_count": mine.filter(Survey.status == SURVEY_VIDEO_PENDING).count(),
        "completed_count": mine.filter(Survey.status == SURVEY_COMPLETED).count(),
    }


def _captain_home(state_scoped):
    from core.engine import active_survey_for

    captain = resolve_actor(roles=("captain", "backup_captain", "admin"))
    active = active_survey_for(captain.email)

    return ok(
        {
            "user": spec("user").serialize(captain, Options.from_request()),
            **_captain_counts(captain, state_scoped),
            "active_survey": (
                {"id": active.id, "status": active.status, "section_no": active.section_no}
                if active
                else None
            ),
            # The portal redirects to /recording whenever a survey is in flight;
            # this encodes that decision for a client to honour.
            "next_screen": "recording" if active else "home",
        }
    )


@bp.get("/dashboards/captain/home")
@require_auth("dashboards:read")
def captain_home():
    return _captain_home(state_scoped=False)


@bp.get("/dashboards/backup/home")
@require_auth("dashboards:read")
def backup_home():
    return _captain_home(state_scoped=True)


@bp.get("/dashboards/captain/assignments")
@require_auth("dashboards:read")
def captain_assignments():
    """The select-stretch picker."""
    from models.db_models import SurveyAssignment

    captain = resolve_actor(roles=("captain", "backup_captain", "admin"))
    query = base_query(SurveyAssignment)

    if _flag("state_scoped") or captain.role == "backup_captain":
        query = query.filter(SurveyAssignment.state == captain.state)
        query = query.order_by(SurveyAssignment.stretch_code)
    else:
        query = query.filter(SurveyAssignment.captain_email == captain.email)

    if _flag("resurvey_only"):
        query = query.filter(SurveyAssignment.survey_enabled.is_(True))
        query = query.order_by(SurveyAssignment.section_no)

    rows = query.all()
    return ok(
        {
            "user": spec("user").serialize(captain, Options.from_request()),
            "assignments": serialize_many(rows, spec("assignment"), Options.from_request()),
        }
    )


@bp.get("/dashboards/captain/checklist")
@require_auth("dashboards:read")
def captain_checklist():
    """The 20 pre-survey items, plus where to submit them."""
    from core.config import as_items, PRE_SURVEY_CHECKLIST

    return ok(
        {
            "items": as_items(PRE_SURVEY_CHECKLIST),
            "requires_dashcam_photo": True,
            "upload_endpoint": "/api/v1/media/images",
            "submit_endpoint": "/api/v1/surveys/start",
        }
    )


@bp.get("/dashboards/captain/recording")
@require_auth("dashboards:read")
def captain_recording():
    """The live-survey screen: the active survey plus the 7 verification items."""
    from core.config import as_items, RECORDING_CHECKLIST
    from core.engine import active_survey_for

    captain = resolve_actor(roles=("captain", "backup_captain", "admin"))
    survey = active_survey_for(captain.email)
    if not survey:
        return ok(
            {
                "survey": None,
                "verification_items": as_items(RECORDING_CHECKLIST),
                "message": "No survey is currently in progress.",
            }
        )

    return ok(
        {
            "survey": spec("survey").serialize(survey, Options.from_request()),
            "derived": survey_derived(survey),
            "verification_items": as_items(RECORDING_CHECKLIST),
            "can_groundwork_complete": survey.status == SURVEY_ONGOING,
            "requires_pdf": True,
            "complete_endpoint": f"/api/v1/surveys/{survey.id}/complete",
        }
    )


@bp.get("/dashboards/captain/pending-uploads")
@require_auth("dashboards:read")
def captain_pending_uploads():
    """
    Surveys awaiting a video count, or a PDF re-upload.

    routes/captain.py:729-741 - ``status == video_pending OR pdf_reupload_required``.
    """
    from extensions import db
    from models.db_models import Survey

    captain = resolve_actor(roles=("captain", "backup_captain", "admin"))
    rows = (
        base_query(Survey)
        .filter(
            Survey.captain_email == captain.email,
            db.or_(
                Survey.status == SURVEY_VIDEO_PENDING,
                Survey.pdf_reupload_required.is_(True),
            ),
        )
        .order_by(Survey.id.desc())
        .all()
    )
    return ok({"surveys": _rows(rows)})


@bp.get("/dashboards/backup/pending-uploads")
@require_auth("dashboards:read")
def backup_pending_uploads():
    return captain_pending_uploads()


@bp.get("/dashboards/captain/completed-surveys")
@require_auth("dashboards:read")
def captain_completed():
    from models.db_models import Survey

    captain = resolve_actor(roles=("captain", "backup_captain", "admin"))
    rows = (
        base_query(Survey)
        .filter(
            Survey.captain_email == captain.email, Survey.status == SURVEY_COMPLETED
        )
        .order_by(Survey.id.desc())
        .all()
    )
    return ok({"surveys": _rows(rows)})


@bp.get("/dashboards/backup/completed-surveys")
@require_auth("dashboards:read")
def backup_completed():
    return captain_completed()


@bp.get("/dashboards/captain/video-counts/<int:survey_id>")
@require_auth("dashboards:read")
def captain_video_counts(survey_id):
    """The 8 numeric inputs, with their current values."""
    from models.db_models import Survey

    from core.config import DEFECT_FIELD_META

    survey = get_or_404(Survey, survey_id, wants_deleted())
    return ok(
        {
            "survey": spec("survey").serialize(survey, Options.from_request()),
            "fields": [
                {
                    "name": field,
                    "side": side,
                    "label": label,
                    "current_value": getattr(survey, field, 0) or 0,
                }
                for field, (side, label) in DEFECT_FIELD_META.items()
            ],
            "submit_endpoint": f"/api/v1/surveys/{survey.id}/video-counts",
        }
    )


@bp.get("/dashboards/captain/unable-to-survey")
@require_auth("dashboards:read")
def captain_unable():
    from models.db_models import SurveyAssignment

    captain = resolve_actor(roles=("captain", "backup_captain", "admin"))
    rows = (
        base_query(SurveyAssignment)
        .filter(SurveyAssignment.captain_email == captain.email)
        .all()
    )
    return ok(
        {
            "assignments": serialize_many(rows, spec("assignment"), Options.from_request())
        }
    )

