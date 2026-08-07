from flask import Blueprint
from flask import render_template
from flask import request
from extensions import db
from models.db_models import SurveyAssignment
from models.db_models import User
from flask import session
import pytz
from sqlalchemy import case
from models.db_models import Survey
from models.db_models import User
from flask import redirect
from datetime import datetime, timedelta
from utils.email_templates import (
    build_subject,
    build_email_body
)

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    session,
    url_for,
    flash,
)

from google_drive import (
    download_file_from_drive,
    create_gmail_draft,
    get_cc_email
)

from utils.request_params import (
    safe_date,
    safe_int,
    safe_week
)

admin_bp = Blueprint(
    "admin",
    __name__
)

from datetime import datetime

@admin_bp.route("/admin")
def admin_dashboard():
    
    if not session.get("user_id"):
     return redirect("/")

    if session.get("role") != "admin":
     return redirect("/")

    from datetime import datetime, timedelta

    ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)

    today = ist_now.strftime("%A")
    current_hour = ist_now.hour

    print("IST TIME:", ist_now)
    print("TODAY:", today)
    print("CURRENT HOUR:", current_hour)

    days_since_monday = ist_now.weekday()

    week_start = (
    ist_now - timedelta(days=days_since_monday)
).replace(
    hour=0,
    minute=0,
    second=0,
    microsecond=0
)

# -----------------------------------
# WEEKLY RESET (Every Monday 12 AM onwards)
# -----------------------------------

    if today == "Monday":

       today_date = ist_now.date()

       reset_required = SurveyAssignment.query.filter(
        db.or_(
            SurveyAssignment.last_week_reset.is_(None),
            SurveyAssignment.last_week_reset < today_date
        )
     ).first()

       if reset_required:

        SurveyAssignment.query.update({

            "status": "assigned",

            "alert_acknowledged": False,

            "last_week_reset": today_date

        })

        db.session.commit()

        # This is a bulk SQL update, so the API's change-capture listeners never
        # see it and the consuming site would silently miss the whole reset.
        # Announce it explicitly. Deliberately after the commit, and best-effort:
        # the Monday reset must not fail because the feed is unavailable.
        try:
            from core.engine import record_bulk_change

            record_bulk_change(
                "assignment",
                reason="weekly Monday reset (routes/admin.py)"
            )
        except Exception:
            pass
    

    total_captains = User.query.filter_by(
        role="captain"
    ).count()

    ongoing = Survey.query.filter_by(
        status="ongoing"
    ).count()

    completed = Survey.query.filter_by(
        status="completed"
    ).count()

    video_pending = Survey.query.filter_by(
        status="video_pending"
    ).count()

    status = request.args.get("status")
    stretch = request.args.get("stretch")
    state = request.args.get("state")
    captain_name = request.args.get("captain_name")
    cycle = request.args.get("cycle")
    week = request.args.get("week")
    day = request.args.get("day")
    agency = request.args.get("agency", "")
    

    query = Survey.query.filter_by(
    show_on_dashboard=True
)
    

    roadvision_pending_query = Survey.query.filter(
    Survey.status == "completed",
    Survey.roadvision_completed.is_(False)
)

    
    if status:
        query = query.filter_by(status=status)

    if stretch:
        query = query.filter(
            Survey.stretch_code.ilike(
                f"%{stretch}%"
            )
        )

    if state:
        query = query.filter_by(state=state)

    if agency == "krish":

     query = query.filter(
        Survey.state.in_([
            "MEGHALAYA",
            "WEST BENGAL",
            "ASSAM",
            "BIHAR"
        ])
    )

    elif agency == "godbole":
 
     query = query.filter(
        Survey.state == "ODISHA"
    )

    elif agency == "aspizo":

     query = query.filter(
        Survey.state.in_([
            "UTTAR PRADESH",
            "JHARKHAND"
        ])
    )

    if captain_name:
        query = query.filter_by(
            captain_name=captain_name
        )

    cycle_no = safe_int(cycle)

    if cycle_no is not None:
        query = query.filter_by(
        cycle_no=cycle_no
    )

    # `day` filters both queries, but only when one was actually chosen. The
    # roadvision line used to sit outside this block, so with no day filter -
    # the default view - it ran as filter_by(survey_day=None) and matched only
    # rows with a NULL survey_day. The RoadVision-pending counter therefore
    # read 0 until an admin picked a day.
    if day:
        query = query.filter_by(
        survey_day=day
    )

        roadvision_pending_query = roadvision_pending_query.filter_by(
            survey_day=day
        )

# -----------------------------------
# PROJECT WEEK FILTER
# -----------------------------------

    week_no = safe_week(week)

    if week_no is not None:

      project_start = datetime(2026, 6, 22)

      start = project_start + timedelta(
        days=(week_no - 1) * 7
      )

      end = start + timedelta(days=7)

      query = query.filter(
        Survey.start_time >= start,
        Survey.start_time < end
      )

      roadvision_pending_query = roadvision_pending_query.filter(
       Survey.start_time >= start,
       Survey.start_time < end
)


    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")

    from_dt = safe_date(from_date)
    to_dt = safe_date(to_date)

    if from_dt:
         query = query.filter(Survey.start_time >= from_dt)

    if to_dt:
         query = query.filter(Survey.start_time < to_dt + timedelta(days=1))

    captains = User.query.filter_by(
        role="captain"
    ).order_by(
        User.name
    ).all()

    states = SurveyAssignment.query.with_entities(
        SurveyAssignment.state
    ).distinct().order_by(
        SurveyAssignment.state
    ).all()

    cycles = db.session.query(
    Survey.cycle_no
    ).distinct().order_by(
    Survey.cycle_no
    ).all()

# -----------------------------
# PROJECT WEEK DROPDOWN
# -----------------------------

    project_start = datetime(2026, 6, 22)

    today_date = datetime.utcnow()

    total_weeks = (
        (today_date.date() - project_start.date()).days // 7
         ) + 1

    weeks = list(range(1, total_weeks + 1))

    from sqlalchemy import case, and_, or_

    all_surveys = query.order_by(

     case(

        (
            Survey.status == "ongoing",
            1
        ),

        (
            Survey.status == "groundwork_completed",
            2
        ),

        (
            Survey.status == "video_uploaded_pending_form",
            3
        ),

        (
            Survey.status == "video_pending",
            4
        ),

        (
            and_(
                Survey.status == "completed",
                or_(
                    Survey.task1_completed == False,
                    Survey.task2_completed == False,
                    Survey.survey_form_completed == False
                )
            ),
            5
        ),

        (
            and_(
                Survey.status == "completed",
                Survey.task1_completed == True,
                Survey.task2_completed == True,
                Survey.survey_form_completed == True
            ),
            6
        ),

        else_=7

    ),

    Survey.start_time.desc()

).all()



    # -----------------------------------
# TEAM KM CALCULATION
# -----------------------------------
    krish_km = 0
    godbole_km = 0
    aspizo_km = 0

    team1_states = {}
    team2_states = {}
    team3_states = {}

    for survey in all_surveys:

       if survey.status != "completed":
        continue

       km = survey.section_length or 0

       if survey.state in [
        "MEGHALAYA",
        "WEST BENGAL",
        "ASSAM",
        "BIHAR"
    ]:

        krish_km += km

        team1_states[survey.state] = (
            team1_states.get(survey.state, 0) + km
        )

       elif survey.state == "ODISHA":

        godbole_km += km

        team2_states[survey.state] = (
            team2_states.get(survey.state, 0) + km
        )

       elif survey.state in [
        "UTTAR PRADESH",
        "JHARKHAND"
    ]:

        aspizo_km += km

        team3_states[survey.state] = (
            team3_states.get(survey.state, 0) + km
        )

    total_km = (
    krish_km +
    godbole_km +
    aspizo_km
)
    resurvey_requests = Survey.query.filter_by(
    resurvey_requested=True,
    resurvey_approved=False
).all()
    

    for survey in all_surveys:

     assignment = SurveyAssignment.query.filter_by(
        section_no=survey.section_no
     ).first()

     if assignment:
        survey.scheduled_day = assignment.survey_day
     else:
        survey.scheduled_day = survey.survey_day


    # -----------------------------------
    # IST TIME DISPLAY
    # -----------------------------------

    for survey in all_surveys:

           if survey.start_time:

            survey.display_start_time = (
                survey.start_time +
                timedelta(hours=5, minutes=30)
            )

           else:

            survey.display_start_time = None

           if survey.end_time:

            survey.display_end_time = (
                survey.end_time +
                timedelta(hours=5, minutes=30)
            )

           else:

            survey.display_end_time = None

            if survey.survey_form_completed:
                survey.survey_form_completed_at_display = (
                    survey.survey_form_completed_at +
                    timedelta(hours=5, minutes=30)
                )

            else:

                survey.survey_form_completed_at_display = None

            if survey.task1_completed:
                survey.task1_completed_at_display = (
                    survey.task1_completed_at +
                    timedelta(hours=5, minutes=30)
                )
            else:

                survey.task1_completed_at_display = None

            if survey.task2_completed:
                survey.task2_completed_at_display = (
                    survey.task2_completed_at +
                    timedelta(hours=5, minutes=30)
                )

            else:

                survey.task2_completed_at_display = None    

           if (
            survey.status == "video_pending"
            and survey.video_pending_start_time
        ):

            survey.upload_duration_minutes = int(
                (
                    datetime.utcnow() -
                    survey.video_pending_start_time
                ).total_seconds() / 60
            )

            survey.upload_status_text = (
                "Upload Pending"
            )

           elif (
            survey.video_pending_start_time
            and survey.video_upload_time
        ):

            survey.upload_duration_minutes = int(
                (
                    survey.video_upload_time -
                    survey.video_pending_start_time
                ).total_seconds() / 60
            )

            survey.upload_status_text = (
                "Upload Duration"
            )

           else:

            survey.upload_duration_minutes = 0

            survey.upload_status_text = ""

    # -----------------------------------
    # MISSED SURVEY LOGIC
    # -----------------------------------

    alerts = []

    today_assignments = SurveyAssignment.query.filter_by(
    survey_day=today,
    survey_enabled=True
    ).all()
    

    for a in today_assignments:
     print(
        a.id,
        a.survey_day,
        a.main_person,
        a.status,
        a.survey_enabled
    )

    if current_hour >= 14:

        captains_today = {}

        for assignment in today_assignments:

            email = assignment.captain_email

            if email not in captains_today:
                captains_today[email] = []

            captains_today[email].append(
                assignment
            )

        for email, surveys in captains_today.items():

            total_surveys = len(
                surveys
            )

            started_surveys = 0

            for assignment in surveys:

                survey_exists = Survey.query.filter(
                Survey.section_no == assignment.section_no,
                Survey.start_time >= week_start,
                Survey.status.in_([
                     "ongoing",
                     "groundwork_completed",
                     "video_pending",
                     "completed"
                                 ])
                     ).first()

                if survey_exists:
             
                        started_surveys += 1

                if survey_exists and assignment.status == "missed":

                 assignment.status = "started"
 
                 assignment.alert_acknowledged = False
 
            if total_surveys == 1:

                if started_surveys == 0:

                    surveys[0].status = "missed"

                    if not surveys[0].alert_acknowledged:

                     alerts.append({

    "assignment_id": surveys[0].id,

    "captain": surveys[0].main_person,

    "message": "Survey not started by 2 PM",

    "state": surveys[0].state,

    "stretch": surveys[0].stretch_code,

    "section_no": surveys[0].section_no

})
            # This block used to sit one level out, OUTSIDE the
            # `for email, surveys in captains_today.items()` loop above. Two
            # things went wrong as a result:
            #
            #   1. With no assignments scheduled today the loop never ran, so
            #      `total_surveys` was never bound and this line raised
            #      UnboundLocalError - a hard 500 on /admin. Assignments only
            #      exist Mon-Fri, so the admin dashboard broke every Saturday
            #      and Sunday after 14:00.
            #
            #   2. When it did run it saw only the LAST captain's leftover
            #      loop variables, so every other captain with two or more
            #      assignments was silently skipped.
            #
            # Indented back into the loop, it now runs once per captain.
            if total_surveys >= 2:

                if started_surveys == 0:

                    for assignment in surveys:

                        assignment.status = "missed"

                        if not assignment.alert_acknowledged:

                            alerts.append({

                                "assignment_id": assignment.id,

                                "captain": assignment.main_person,

                                "message": "Survey not started by 2 PM",

                                "state": assignment.state,

                                "stretch": assignment.stretch_code,

                                "section_no": assignment.section_no

                            })

    missed = SurveyAssignment.query.filter_by(
    status="missed"
    ).count()

    roadvision_pending = roadvision_pending_query.count()

    db.session.commit()

 
    return render_template(
    "admin/dashboard.html",
    total_captains=total_captains,
    ongoing=ongoing,
    completed=completed,
    video_pending=video_pending,
    missed=missed,
    all_surveys=all_surveys,
    states=states,
    captains=captains,
    alerts=alerts,
    cycles=cycles,
    weeks=weeks,
    agency=agency,
    resurvey_requests=resurvey_requests,
    roadvision_pending=roadvision_pending,
    team1_states=team1_states,
    team2_states=team2_states,
    team3_states=team3_states,

    krish_km=round(krish_km, 2),
    godbole_km=round(godbole_km, 2),
    aspizo_km=round(aspizo_km, 2),
    total_km=round(total_km, 2)
)


@admin_bp.route("/admin/missed")
def admin_missed():

    if session.get("role") != "admin":
        return redirect("/")

    agency = request.args.get("agency")
    day = request.args.get("day")

    query = SurveyAssignment.query.filter_by(
        status="missed"
    )

    if agency == "krish":
     query = query.filter(
        SurveyAssignment.state.in_([
            "ASSAM",
            "BIHAR",
            "MEGHALAYA",
            "WEST BENGAL"
        ])
    )

    elif agency == "godbole":
     query = query.filter(
        SurveyAssignment.state == "ODISHA"
    )

    elif agency == "aspizo":
     query = query.filter(
        SurveyAssignment.state.in_([
            "UTTAR PRADESH",
            "JHARKHAND"
        ])
    )

    if day:
        query = query.filter(
            SurveyAssignment.survey_day == day
        )

    missed_surveys = query.order_by(
    SurveyAssignment.survey_day,
    SurveyAssignment.section_no
).all()

# ---------------------------------------
# Calculate Next Cycle Number
# ---------------------------------------

    for assignment in missed_surveys:

     latest = (
        Survey.query.filter_by(
            section_no=assignment.section_no
        )
        .order_by(Survey.cycle_no.desc())
        .first()
    )

    assignment.next_cycle = (
        latest.cycle_no + 1
        if latest
        else 1
    )

    return render_template(
    "admin/missed.html",
    missed_surveys=missed_surveys
)


@admin_bp.route("/admin/missed/extract")
def missed_extract():

    if session.get("role") != "admin":
        return redirect("/")

    agency = request.args.get("agency")
    day = request.args.get("day")

    query = SurveyAssignment.query.filter_by(
        status="missed"
    )

    if agency == "krish":
     query = query.filter(
        SurveyAssignment.state.in_([
            "ASSAM",
            "BIHAR",
            "MEGHALAYA",
            "WEST BENGAL"
        ])
    )

    elif agency == "godbole":
     query = query.filter(
        SurveyAssignment.state == "ODISHA"
    )

    elif agency == "aspizo":
     query = query.filter(
        SurveyAssignment.state.in_([
            "UTTAR PRADESH",
            "JHARKHAND"
        ])
    )

    if day:
        query = query.filter(
            SurveyAssignment.survey_day == day
        )

    missed_surveys = query.order_by(
        SurveyAssignment.section_no
    ).all()

    missed_list = []

    for assignment in missed_surveys:

        latest_survey = Survey.query.filter_by(
            section_no=assignment.section_no
        ).order_by(
            Survey.cycle_no.desc()
        ).first()

        if latest_survey:
            next_cycle = latest_survey.cycle_no + 1
        else:
            next_cycle = 1

        missed_list.append({
            "section_no": assignment.section_no,
            "cycle_no": next_cycle
        })

    return render_template(
        "admin/missed_extract.html",
        missed_list=missed_list
    )

@admin_bp.route("/admin/survey/<int:survey_id>")
def survey_details_admin(survey_id):
    if session.get("role") != "admin":
     return redirect("/")

    from datetime import timedelta

    survey = Survey.query.get_or_404(
        survey_id
    )

    display_start_time = None
    display_end_time = None

    if survey.start_time:

        display_start_time = (
            survey.start_time +
            timedelta(hours=5, minutes=30)
        )

    if survey.end_time:

        display_end_time = (
            survey.end_time +
            timedelta(hours=5, minutes=30)
        )

    return render_template(
        "admin/survey_details.html",
        survey=survey,
        display_start_time=display_start_time,
        display_end_time=display_end_time
    )


@admin_bp.route(
    "/acknowledge-alert/<int:assignment_id>"
)
def acknowledge_alert(
    assignment_id
):
    if session.get("role") != "admin":
     return redirect("/")

    assignment = SurveyAssignment.query.get_or_404(
        assignment_id
    )

    # Was a bulk .update(), which bypasses the ORM and so emitted no webhook and
    # no change-log row. This touches a handful of rows for one captain on one
    # day, so iterating costs nothing and every row now produces a proper
    # assignment.alert_acknowledged event.
    siblings = SurveyAssignment.query.filter_by(
        captain_email=assignment.captain_email,
        survey_day=assignment.survey_day
    ).all()

    for sibling in siblings:
        sibling.alert_acknowledged = True

    db.session.commit()

    return redirect("/admin")


@admin_bp.route(
    "/approve-resurvey/<int:survey_id>"
)
def approve_resurvey(survey_id):

    if session.get("role") != "admin":
        return redirect("/")

    survey = Survey.query.get_or_404(
        survey_id
    )

    survey.resurvey_approved = True

    db.session.commit()

    return redirect("/admin")


@admin_bp.route("/admin/schedules")
def admin_schedules():

    if session.get("role") != "admin":
        return redirect("/")

    day = request.args.get("day")
    state = request.args.get("state")

    query = SurveyAssignment.query

    if day:
        query = query.filter_by(
            survey_day=day
        )

    if state:
        query = query.filter_by(
            state=state
        )

    from sqlalchemy import case

    day_order = case(
    (SurveyAssignment.survey_day == "Monday", 1),
    (SurveyAssignment.survey_day == "Tuesday", 2),
    (SurveyAssignment.survey_day == "Wednesday", 3),
    (SurveyAssignment.survey_day == "Thursday", 4),
    (SurveyAssignment.survey_day == "Friday", 5),
    else_=6
)

    schedules = query.order_by(
    day_order,
    SurveyAssignment.state,
    SurveyAssignment.section_no
).all()

    states = SurveyAssignment.query.with_entities(
        SurveyAssignment.state
    ).distinct().all()

    monday_count = SurveyAssignment.query.filter_by(
        survey_day="Monday"
    ).count()

    tuesday_count = SurveyAssignment.query.filter_by(
        survey_day="Tuesday"
    ).count()

    wednesday_count = SurveyAssignment.query.filter_by(
        survey_day="Wednesday"
    ).count()

    thursday_count = SurveyAssignment.query.filter_by(
        survey_day="Thursday"
    ).count()

    friday_count = SurveyAssignment.query.filter_by(
        survey_day="Friday"
    ).count()

    return render_template(
        "admin/schedules.html",
        schedules=schedules,
        states=states,
        monday_count=monday_count,
        tuesday_count=tuesday_count,
        wednesday_count=wednesday_count,
        thursday_count=thursday_count,
        friday_count=friday_count
    )


@admin_bp.route(
    "/request-pdf-reupload/<int:survey_id>",
    methods=["POST"]
)
def request_pdf_reupload(survey_id):

    if session.get("role") != "admin":
        return redirect("/")

    survey = Survey.query.get_or_404(survey_id)

    survey.pdf_reupload_required = True

    survey.pdf_reupload_reason = request.form["reason"]

    survey.pdf_reupload_count += 1

    db.session.commit()

    return redirect(f"/admin/survey/{survey.id}")


@admin_bp.route("/reports")
def reports():

    if not session.get("user_id"):
        return redirect("/")

    if session.get("role") != "admin":
        return redirect("/")

    from datetime import datetime, timedelta

    query = Survey.query.filter_by(
        status="completed"
    )

    # --------------------------
    # FILTERS
    # --------------------------

    week = request.args.get("week")
    cycle = request.args.get("cycle")
    team = request.args.get("team")
    state = request.args.get("state")
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")

    # Cycle

    cycle_no = safe_int(cycle)

    if cycle_no is not None:
        query = query.filter_by(
            cycle_no=cycle_no
        )

    # Week

    project_start = datetime(2026, 6, 22)

    week_no = safe_week(week)

    if week_no is not None:

        start = project_start + timedelta(
            days=(week_no - 1) * 7
        )

        end = start + timedelta(days=7)

        query = query.filter(
            Survey.start_time >= start,
            Survey.start_time < end
        )

    # Date

    from_dt = safe_date(from_date)
    to_dt = safe_date(to_date)

    if from_dt:

        query = query.filter(
            Survey.start_time >= from_dt
        )

    if to_dt:

        query = query.filter(
            Survey.start_time <
            to_dt + timedelta(days=1)
        )

    # Team

    if team == "Krish":

        query = query.filter(
            Survey.state.in_([
                "WEST BENGAL",
                "ASSAM",
                "BIHAR"
            ])
        )

    elif team == "Godbole":

        query = query.filter(
            Survey.state == "ODISHA"
        )

    elif team == "Aspizo":

        query = query.filter(
            Survey.state.in_([
                "UTTAR PRADESH",
                "JHARKHAND"
            ])
        )

    # State

    if state:
         query = query.filter(
        Survey.state == state
    )

    surveys = query.all()

    # --------------------------
    # SUMMARY
    # --------------------------

    completed_surveys = len(surveys)

    total_km = sum(
        s.section_length or 0
        for s in surveys
    )

    captains = len(set(
        s.captain_email
        for s in surveys
    ))

    total_minutes = 0

    for survey in surveys:

        if survey.start_time and survey.end_time:

            total_minutes += (
                survey.end_time -
                survey.start_time
            ).total_seconds() / 60

    total_hours = round(
        total_minutes / 60,
        2
    )

    # --------------------------
    # FILTERS
    # --------------------------

    today = datetime.utcnow()

    total_weeks = (
        (today.date()-project_start.date()).days//7
    ) + 1

    weeks = list(
        range(
            1,
            total_weeks+1
        )
    )

    cycles = db.session.query(
        Survey.cycle_no
    ).distinct().order_by(
        Survey.cycle_no
    ).all()

    return render_template(

        "admin/reports.html",

        weeks=weeks,
        cycles=cycles,
        surveys = surveys,
        completed_surveys=completed_surveys,
        total_km=round(total_km,2),
        captains=captains,
        total_hours=total_hours
    )


@admin_bp.route("/reports/export")
def export_report():

    if not session.get("user_id"):
        return redirect("/")

    if session.get("role") != "admin":
        return redirect("/")

    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font

    # -------------------------
    # SAME FILTERS AS REPORT PAGE
    # -------------------------

    query = Survey.query.filter_by(
        status="completed"
    )

    week = request.args.get("week")
    cycle = request.args.get("cycle")
    team = request.args.get("team")
    state = request.args.get("state")
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")
    print("Selected State:", state)

    from datetime import datetime, timedelta

    cycle_no = safe_int(cycle)

    if cycle_no is not None:
        query = query.filter_by(
            cycle_no=cycle_no
        )

    project_start = datetime(2026, 6, 22)

    week_no = safe_week(week)

    if week_no is not None:

        start = project_start + timedelta(
            days=(week_no - 1) * 7
        )

        end = start + timedelta(days=7)

        query = query.filter(
            Survey.start_time >= start,
            Survey.start_time < end
        )

    from_dt = safe_date(from_date)
    to_dt = safe_date(to_date)

    if from_dt:

        query = query.filter(
            Survey.start_time >= from_dt
        )

    if to_dt:

        query = query.filter(
            Survey.start_time <
            to_dt + timedelta(days=1)
        )

    if team == "Krish":

        query = query.filter(
            Survey.state.in_([
                "WEST BENGAL",
                "ASSAM",
                "BIHAR"
            ])
        )

    elif team == "Godbole":

        query = query.filter(
            Survey.state == "ODISHA"
        )

    elif team == "Aspizo":

        query = query.filter(
            Survey.state.in_([
                "UTTAR PRADESH",
                "JHARKHAND"
            ])
        )

    # State

    if state:
        query = query.filter(
        Survey.state == state
    )

    surveys = query.order_by(
        Survey.start_time.desc()
    ).all()

    # -------------------------
    # EXCEL
    # -------------------------

    wb = Workbook()

    ws = wb.active

    ws.title = "Survey Report"

    headers = [

        "Date",
        "Cycle",
        "Captain",
        "State",
        "Section No",
        "UPC Code",
        "Stretch",
        "KM",
        "Survey Type",
        "Status"

    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=1,
            column=col
        )

        cell.value = header
        cell.font = Font(bold=True)

    row = 2

    for survey in surveys:

        ws.cell(
            row=row,
            column=1
        ).value = (
            survey.start_time.strftime("%d-%m-%Y")
            if survey.start_time
            else ""
        )

        ws.cell(row=row,column=2).value = survey.cycle_no
        ws.cell(row=row,column=3).value = survey.captain_name
        ws.cell(row=row,column=4).value = survey.state
        ws.cell(row=row,column=5).value = survey.section_no
        ws.cell(row=row,column=6).value = survey.upc_code
        ws.cell(row=row,column=7).value = survey.stretch_code
        ws.cell(row=row,column=8).value = survey.section_length
        ws.cell(row=row,column=9).value = survey.survey_type
        ws.cell(row=row,column=10).value = survey.status

        row += 1

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name="Survey_Report.xlsx",

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )


@admin_bp.route("/admin/gmail-drafts")
def gmail_drafts():

    if not session.get("user_id"):
        return redirect("/")

    if session.get("role") != "admin":
        return redirect("/")

    return render_template(
        "admin/gmail_drafts.html"
    )

@admin_bp.route(
    "/admin/gmail-drafts/<email_type>",
    methods=["GET", "POST"]
)
def email_draft(email_type):

    if not session.get("user_id"):
        return redirect("/")

    if session.get("role") != "admin":
        return redirect("/")

    if email_type not in ["defect", "raw", "discrepancy"]:
        return redirect("/admin/gmail-drafts")

    project_start = datetime(2026, 6, 22)
    today = datetime.utcnow()

    total_weeks = (
        (today.date() - project_start.date()).days // 7
    ) + 1

    weeks = list(range(1, total_weeks + 1))

    cycles = (
        db.session.query(Survey.cycle_no)
        .distinct()
        .order_by(Survey.cycle_no)
        .all()
    )

    surveys = []
    no_records = False

    selected_week = request.form.get("week") or request.args.get("week", "")
    selected_section = request.form.get("survey_id", "")
    upc_code = request.form.get("upc_code", "").strip()
    cycle_no = request.form.get("cycle_no", "")

    start_date = request.form.get("start_date", "")
    end_date = request.form.get("end_date", "")

    subject = None
    email_html = None

    # --------------------------------------------------
    # LOAD SURVEYS
    # --------------------------------------------------

    query = Survey.query.filter(
        Survey.status.in_([
            "ongoing",
            "video_uploaded_pending_form",
            "groundwork_completed",
            "video_pending",
            "completed",
        ])
    )

    if selected_week:
        week_no = int(selected_week)

        start = project_start + timedelta(days=(week_no - 1) * 7)
        end = start + timedelta(days=7)

        query = query.filter(
            Survey.start_time >= start,
            Survey.start_time < end,
        )

    if upc_code:
        query = query.filter(
            Survey.upc_code.ilike(f"%{upc_code}%")
        )

    if cycle_no:
        query = query.filter(
            Survey.cycle_no == int(cycle_no)
        )

    surveys = (
        query.order_by(
            Survey.section_no,
            Survey.cycle_no,
        )
        .all()
    )

    # --------------------------------------------------
    # ADD MISSED SURVEYS
    # --------------------------------------------------

    missed_query = SurveyAssignment.query.filter_by(
        status="missed"
    )

    if upc_code:
        missed_query = missed_query.filter(
            SurveyAssignment.upc_code.ilike(f"%{upc_code}%")
        )

    missed_assignments = missed_query.all()

    for assignment in missed_assignments:

        latest = (
            Survey.query.filter_by(
                section_no=assignment.section_no
            )
            .order_by(Survey.cycle_no.desc())
            .first()
        )

        assignment.missed_id = f"missed-{assignment.id}"
        assignment.cycle_no = (
            latest.cycle_no + 1 if latest else 1
        )
        assignment.is_missed = True

        surveys.append(assignment)

    surveys.sort(
        key=lambda x: (
            str(x.section_no),
            int(x.cycle_no),
        )
    )

    if (
        request.method == "POST"
        and request.form.get("action") != "generate"
    ):
        if not surveys:
            no_records = True

    # --------------------------------------------------
    # GENERATE EMAIL
    # --------------------------------------------------

    if (
        request.method == "POST"
        and request.form.get("action") == "generate"
        and selected_section
    ):

        is_missed = False

        if selected_section.startswith("missed-"):

            is_missed = True

            assignment_id = int(
                selected_section.replace("missed-", "")
            )

            assignment = SurveyAssignment.query.get_or_404(
                assignment_id
            )

            survey = (
             Survey.query.filter_by(
             section_no=assignment.section_no
    )
             .order_by(Survey.cycle_no.desc())
             .first()
)

            if survey is None:
               flash(
               f"No previous survey found for Section {assignment.section_no}. "
               "A Gmail draft cannot be generated.",
               "error",
    )
               return redirect(
               url_for(
            "admin.email_draft",
            email_type=email_type,
            week=selected_week,
        )
    )

        else:

            survey = Survey.query.get_or_404(
                safe_int(selected_section, minimum=1) or 0
            )

        original_cycle = survey.cycle_no

        if is_missed:
            survey.cycle_no = original_cycle + 1

        # --------------------------------------------------
        # Validation
        # --------------------------------------------------

        # Your validation code here

        # --------------------------------------------------
        # Build Email
        # --------------------------------------------------

        subject = build_subject(
            survey,
            email_type,
            end_date,
        )

        subject = " ".join(str(subject).split())

        email_html = build_email_body(
            survey,
            email_type,
            start_date,
            end_date,
        )

        # --------------------------------------------------
        # Attachment
        # --------------------------------------------------

        if email_type == "defect":

            attachment_url = survey.defect_report_file

            attachment_name = (
                f"{survey.upc_code}_Cycle-{survey.cycle_no}_Defect_Report.xlsx"
            )

        elif email_type == "raw":

            attachment_url = survey.raw_video_excel_file

            attachment_name = (
                f"{survey.upc_code}_Cycle-{survey.cycle_no}_Raw_Data.xlsx"
            )

        elif email_type == "discrepancy":

            attachment_url = ""
            attachment_name = ""

        attachment_bytes = None

        if attachment_url:
            attachment_bytes = download_file_from_drive(
                attachment_url
            )

        # --------------------------------------------------
        # Create Gmail Draft
        # --------------------------------------------------

        cc_email = get_cc_email(survey.ro)

        draft_id = create_gmail_draft(
            subject=subject,
            html_body=email_html,
            attachment_bytes=attachment_bytes,
            attachment_filename=attachment_name,
            cc_email=cc_email,
        )

        if is_missed:
            survey.cycle_no = original_cycle

        # --------------------------------------------------
        # Open Gmail Draft
        # --------------------------------------------------



        flash(
            (
                f"{survey.section_no}_Cycle{survey.cycle_no} "
                f"Gmail draft created successfully.\n"
                f"The draft has been created in "
                f"adordashcam@gmail.com. "
                f"Please open Gmail → Drafts and send the email."
            ),
            "success",
        )

        return redirect(
            url_for(
                "admin.email_draft",
                email_type=email_type,
                week=selected_week,
            )
        )

    return render_template(
        "admin/email_draft.html",
        email_type=email_type,
        weeks=weeks,
        cycles=cycles,
        surveys=surveys,
        no_records=no_records,
        selected_week=selected_week,
        selected_section=selected_section,
        upc_code=upc_code,
        cycle_no=cycle_no,
        start_date=start_date,
        end_date=end_date,
        subject=subject,
        email_html=email_html,
    )


@admin_bp.route("/admin/remark/<int:survey_id>")
def admin_view_remark(survey_id):

    if session.get("role") != "admin":
        return redirect("/")

    survey = Survey.query.get_or_404(survey_id)

    return render_template(
        "admin/view_remark.html",
        survey=survey
    )


@admin_bp.route("/admin/extract-list")
def extract_list():

    if session.get("role") != "admin":
        return redirect("/")

    status = request.args.get("status")
    stretch = request.args.get("stretch")
    state = request.args.get("state")
    captain_name = request.args.get("captain_name")
    cycle = request.args.get("cycle")
    week = request.args.get("week")
    day = request.args.get("day")
    agency = request.args.get("agency")
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")

    query = Survey.query.filter_by(
        show_on_dashboard=True
    )

    if status:
        query = query.filter_by(status=status)

    if stretch:
        query = query.filter(
            Survey.stretch_code.ilike(f"%{stretch}%")
        )

    if state:
        query = query.filter_by(state=state)

    if agency == "krish":

     query = query.filter(
        Survey.state.in_([
            "MEGHALAYA",
            "WEST BENGAL",
            "ASSAM",
            "BIHAR"
        ])
    )

    elif agency == "godbole":

     query = query.filter(
        Survey.state == "ODISHA"
    )

    elif agency == "aspizo":

     query = query.filter(
        Survey.state.in_([
            "UTTAR PRADESH",
            "JHARKHAND"
        ])
    )

    if captain_name:
        query = query.filter_by(
            captain_name=captain_name
        )

    cycle_no = safe_int(cycle)

    if cycle_no is not None:
        query = query.filter_by(cycle_no=cycle_no)

    if day:
        query = query.filter_by(survey_day=day)

    project_start = datetime(2026, 6, 22)

    week_no = safe_week(week)

    if week_no is not None:

        start = project_start + timedelta(
            days=(week_no-1)*7
        )

        end = start + timedelta(days=7)

        query = query.filter(
            Survey.start_time >= start,
            Survey.start_time < end
        )

    from_dt = safe_date(from_date)
    to_dt = safe_date(to_date)

    if from_dt:
        query = query.filter(Survey.start_time >= from_dt)

    if to_dt:
        query = query.filter(
            Survey.start_time < to_dt + timedelta(days=1)
        )

    surveys = query.order_by(
        Survey.section_no,
        Survey.cycle_no
    ).all()

    return render_template(
        "admin/extract_list.html",
        surveys=surveys
    )

